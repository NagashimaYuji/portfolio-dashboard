# -*- coding: utf-8 -*-
"""
make_brand_summary.py
最新の資産クラスシートから 投資信託(F) / 日本株(G) / 米国株(H) を
銘柄ごとに株数合算し、時価総額・購入総額・含み損益・含み損益%・
直近為替・購入時為替・カテゴリ内構成比 をまとめた Excel を出力する。
"""

import os, sys, re, glob
from datetime import date, datetime

if sys.stdout.encoding and sys.stdout.encoding.lower() in ('cp932', 'shift_jis', 'shift-jis'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter

BASE_DIR = r'C:\Users\yna78\OneDrive - Corazon\Yuji Private\資産ポートフォリオ戦略'

def find_latest_excel(base_dir):
    files = glob.glob(os.path.join(base_dir, '資産クラス整理_????????.xlsx'))
    if not files:
        raise FileNotFoundError(f'資産クラス整理_YYYYMMDD.xlsx が見つかりません: {base_dir}')
    return max(files)

EXCEL_FILE = find_latest_excel(BASE_DIR)
wb_f = openpyxl.load_workbook(EXCEL_FILE)
wb_c = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

_sn = sorted([s for s in wb_f.sheetnames
              if s.startswith('資産クラス') and s[5:].isdigit() and len(s[5:]) == 8],
             reverse=True)
SN_SHEET = _sn[0]
ws_f = wb_f[SN_SHEET]
ws_c = wb_c[SN_SHEET]
SNAP_DATE = SN_SHEET[5:]

print(f'対象ファイル: {EXCEL_FILE}')
print(f'対象シート  : {SN_SHEET}')

# ── セル評価 (数式チェーン / SUM / 四則演算を解決) ────────────
def ev(r, c, depth=8):
    """キャッシュ優先 → 数式を再帰評価して数値を返す"""
    if depth <= 0:
        return None
    v = ws_c.cell(row=r, column=c).value
    if isinstance(v, (int, float)):
        return float(v)
    v = ws_f.cell(row=r, column=c).value
    if v is None:
        return None
    if isinstance(v, datetime):
        from openpyxl.utils.datetime import to_excel
        try:
            return float(to_excel(v))
        except Exception:
            return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s.startswith('='):
        try:
            return float(s)
        except ValueError:
            return None
    expr = s[1:].replace('$', '').strip()
    # セル参照を値に置換してから評価
    def _sub(m):
        val = ev(int(m.group(2)), column_index_from_string(m.group(1).upper()), depth - 1)
        return f'({val})' if val is not None else 'None'
    expr2 = re.sub(r'\b([A-Z]{1,2})(\d{1,4})\b', _sub, expr)
    if 'None' in expr2:
        return None
    if re.fullmatch(r'[\d\s\+\-\*\/\(\)\.eE]+', expr2):
        try:
            return float(eval(expr2))  # noqa: S307 — 数字と演算子のみ
        except Exception:
            return None
    return None

def sv(r, c, depth=6):
    """文字列セルを取得 (=F117 のような参照も解決)"""
    v = ws_f.cell(row=r, column=c).value
    if v is None:
        return ''
    s = str(v).strip()
    if s.startswith('=') and depth > 0:
        m = re.fullmatch(r'=\$?([A-Z]{1,2})\$?(\d{1,4})', s)
        if m:
            return sv(int(m.group(2)), column_index_from_string(m.group(1).upper()), depth - 1)
    return s

USDJPY = ev(29, 8) or 0.0
print(f'直近為替 USDJPY = {USDJPY}')

# ── 銘柄名の正規化 (表記ゆれ吸収) ────────────────────────────
BRAND_ALIAS = {
    'APPL': 'AAPL',
    'GOLDMANSACS': 'Goldman Sachs',
    'GOLDMAN SACS': 'Goldman Sachs',
    'META': 'Meta',
    'MORGAN STANLEY': 'Morgan Stanley',
}

def norm_brand(name):
    key = name.strip().upper()
    return BRAND_ALIAS.get(key, name.strip())

# ── セクションのデータ行を検出 ────────────────────────────────
def scan_section(label):
    """B列=label のセクションのデータ行リストを返す"""
    in_sec, rows = False, []
    for r in range(1, ws_f.max_row + 1):
        b  = str(ws_f.cell(row=r, column=2).value or '').strip()
        c3 = str(ws_f.cell(row=r, column=3).value or '').strip()
        if b == label:
            in_sec = True
            d = ws_f.cell(row=r, column=4).value
            if d and str(d).strip() not in ('名前', ''):
                rows.append(r)
            continue
        if not in_sec:
            continue
        if b and len(b) == 1 and b.isalpha():
            break
        if c3 == '小計':
            break
        d = ws_f.cell(row=r, column=4).value
        if d and str(d).strip() not in ('名前', ''):
            rows.append(r)
    return rows

# ── 銘柄別集計 ────────────────────────────────────────────────
def aggregate(section, rows):
    """
    銘柄ごとに数量を合算し、時価総額・購入総額等を集計。
      col6=銘柄, col7=評価/時価単価, col9=購入単価, col10=数量
      col14 に =$H$29 があれば USD 建て, col15=購入時為替
    """
    agg = {}
    for r in rows:
        brand = norm_brand(sv(r, 6))
        if not brand:
            continue
        price = ev(r, 7)     # 現在単価
        c_unit = ev(r, 9)    # 購入単価
        qty = ev(r, 10)      # 数量
        if qty is None:
            qty = 0.0
        # col14 が =$H$29 → USD 建て
        n_raw = ws_f.cell(row=r, column=14).value
        is_usd = isinstance(n_raw, str) and 'H29' in n_raw.replace('$', '')
        p_fx = ev(r, 15) if is_usd else None

        a = agg.setdefault(brand, {
            'brand': brand, 'is_usd': is_usd, 'qty': 0.0,
            'price': None, 'cost_jpy': 0.0, 'cost_usd': 0.0,
            'fx_num': 0.0, 'fx_den': 0.0, 'rows': [], 'holders': set(),
        })
        a['is_usd'] = a['is_usd'] or is_usd
        a['qty'] += qty
        a['rows'].append(r)
        holder = sv(r, 4)
        if holder:
            a['holders'].add('Akatsuki' if holder.lower().startswith('aka') else holder)
        if price is not None and a['price'] is None:
            a['price'] = price
        # 購入総額: 行ごとの購入単価 × 数量 ×(USD なら購入時為替)
        if c_unit is not None and qty:
            cost_ccy = c_unit * qty                    # 現地通貨ベース
            rate = (p_fx or USDJPY) if is_usd else 1.0
            a['cost_jpy'] += cost_ccy * rate
            if is_usd:
                a['cost_usd'] += cost_ccy
                a['fx_num'] += rate * cost_ccy         # 購入額加重平均用
                a['fx_den'] += cost_ccy

    out = []
    for a in agg.values():
        price, qty = a['price'], a['qty']
        is_usd = a['is_usd']
        mv = price * qty * (USDJPY if is_usd else 1.0) if (price is not None and qty) else 0.0
        cost = a['cost_jpy']
        gain = mv - cost
        gain_pct = (gain / cost) if cost else None
        avg_fx = (a['fx_num'] / a['fx_den']) if a['fx_den'] else None
        out.append({
            'brand': a['brand'],
            'holders': '/'.join(sorted(a['holders'])),
            'currency': 'US$' if is_usd else 'JPY',
            'qty': qty,
            'price': price,
            'market_value': mv,
            'cost_total': cost,
            'gain': gain,
            'gain_pct': gain_pct,
            'fx_now': USDJPY if is_usd else None,
            'fx_buy': avg_fx,
            'nrows': len(a['rows']),
        })
    out.sort(key=lambda x: -x['market_value'])
    total_mv = sum(x['market_value'] for x in out)
    for x in out:
        x['share'] = (x['market_value'] / total_mv) if total_mv else None
    return out, total_mv

SECTIONS = [('F', '投資信託'), ('G', '日本株'), ('H', '米国株')]
results = {}
for key, label in SECTIONS:
    rows = scan_section(key)
    data, total = aggregate(key, rows)
    results[key] = {'label': label, 'data': data, 'total': total}
    print(f'{key} {label}: {len(rows)}行 → {len(data)}銘柄, 時価合計 {total:,.0f}')

# ══════════════════════════════════════════════════════════════
# Excel 出力
# ══════════════════════════════════════════════════════════════
OUT_FILE = os.path.join(BASE_DIR, f'銘柄別集計_{SNAP_DATE}.xlsx')

wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

HDR_FILL   = PatternFill('solid', fgColor='1F3864')
HDR_FONT   = Font(color='FFFFFF', bold=True, size=10)
TOT_FILL   = PatternFill('solid', fgColor='D9E1F2')
TITLE_FONT = Font(bold=True, size=13, color='1F3864')
THIN = Side(style='thin', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADS = ['銘柄', '保有者', '通貨', '株数/口数', '現在単価', '時価総額(円)',
         '購入総額(円)', '含み損益(円)', '含み損益%', 'カテゴリ構成比%',
         '直近為替', '購入時為替(加重平均)']
WIDTHS = [26, 16, 7, 13, 12, 15, 15, 15, 11, 13, 10, 17]

def style_sheet(ws, title, start_row=1):
    ws.cell(row=start_row, column=1, value=title).font = TITLE_FONT
    hr = start_row + 1
    for i, h in enumerate(HEADS, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BORDER
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return hr

def write_rows(ws, hr, data, total_mv):
    r = hr + 1
    for d in data:
        vals = [d['brand'], d['holders'], d['currency'], d['qty'], d['price'],
                d['market_value'], d['cost_total'], d['gain'], d['gain_pct'],
                d['share'], d['fx_now'], d['fx_buy']]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER
            if i == 4:   c.number_format = '#,##0.####'
            elif i == 5: c.number_format = '#,##0.00##'
            elif i in (6, 7, 8): c.number_format = '#,##0'
            elif i in (9, 10):   c.number_format = '0.00%'
            elif i in (11, 12):  c.number_format = '#,##0.00'
            if i == 8 and d['gain'] is not None:
                c.font = Font(color='006100' if d['gain'] >= 0 else 'C00000', bold=True)
            if i == 9 and d['gain_pct'] is not None:
                c.font = Font(color='006100' if d['gain_pct'] >= 0 else 'C00000')
        r += 1
    # 合計行
    t_cost = sum(x['cost_total'] for x in data)
    t_gain = total_mv - t_cost
    tot = ['合計', '', '', '', '', total_mv, t_cost, t_gain,
           (t_gain / t_cost) if t_cost else None, 1.0 if total_mv else None, '', '']
    for i, v in enumerate(tot, start=1):
        c = ws.cell(row=r, column=i, value=v if v != '' else None)
        c.fill, c.border = TOT_FILL, BORDER
        c.font = Font(bold=True, color='006100' if (i == 8 and t_gain >= 0)
                      else ('C00000' if i == 8 else '000000'))
        if i in (6, 7, 8): c.number_format = '#,##0'
        elif i in (9, 10): c.number_format = '0.00%'
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    return r

# ── 各カテゴリのシート ────────────────────────────────────────
for key, label in SECTIONS:
    res = results[key]
    ws = wb_out.create_sheet(f'{key}_{label}')
    hr = style_sheet(ws, f'{key}. {label} — 銘柄別集計  (基準日: {SNAP_DATE}  USD/JPY: {USDJPY:,.3f})')
    write_rows(ws, hr, res['data'], res['total'])

# ── 全体サマリーシート ────────────────────────────────────────
ws = wb_out.create_sheet('サマリー', 0)
ws.cell(row=1, column=1, value=f'銘柄別集計 サマリー  (基準日: {SNAP_DATE})').font = Font(bold=True, size=14, color='1F3864')
ws.cell(row=2, column=1, value=f'元ファイル: {os.path.basename(EXCEL_FILE)} / シート: {SN_SHEET}').font = Font(size=9, color='666666')
ws.cell(row=3, column=1, value=f'直近為替 USD/JPY: {USDJPY:,.3f}').font = Font(size=9, color='666666')

sh = ['カテゴリ', '銘柄数', '時価総額(円)', '購入総額(円)', '含み損益(円)', '含み損益%', '全体構成比%']
for i, h in enumerate(sh, start=1):
    c = ws.cell(row=5, column=i, value=h)
    c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BORDER
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for i, w in enumerate([20, 9, 16, 16, 16, 12, 13], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

grand_mv = sum(results[k]['total'] for k, _ in SECTIONS)
r = 6
for key, label in SECTIONS:
    res = results[key]
    cost = sum(x['cost_total'] for x in res['data'])
    gain = res['total'] - cost
    vals = [f'{key}. {label}', len(res['data']), res['total'], cost, gain,
            (gain / cost) if cost else None,
            (res['total'] / grand_mv) if grand_mv else None]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = BORDER
        if i in (3, 4, 5): c.number_format = '#,##0'
        elif i in (6, 7):  c.number_format = '0.00%'
        if i == 5 and gain is not None:
            c.font = Font(color='006100' if gain >= 0 else 'C00000', bold=True)
    r += 1

g_cost = sum(sum(x['cost_total'] for x in results[k]['data']) for k, _ in SECTIONS)
g_gain = grand_mv - g_cost
tot = ['3カテゴリ合計', sum(len(results[k]['data']) for k, _ in SECTIONS),
       grand_mv, g_cost, g_gain, (g_gain / g_cost) if g_cost else None, 1.0]
for i, v in enumerate(tot, start=1):
    c = ws.cell(row=r, column=i, value=v)
    c.fill, c.border = TOT_FILL, BORDER
    c.font = Font(bold=True, color='006100' if (i == 5 and g_gain >= 0)
                  else ('C00000' if i == 5 else '000000'))
    if i in (3, 4, 5): c.number_format = '#,##0'
    elif i in (6, 7):  c.number_format = '0.00%'

wb_out.save(OUT_FILE)
print(f'\n[OK] 保存完了: {OUT_FILE}')
print(f'  3カテゴリ合計 時価 {grand_mv:,.0f} / 購入 {g_cost:,.0f} / 損益 {g_gain:,.0f} ({g_gain/g_cost*100:+.2f}%)')
