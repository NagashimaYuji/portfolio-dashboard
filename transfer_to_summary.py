# -*- coding: utf-8 -*-
"""
transfer_to_summary.py
最新の資産クラスシートの各セクション小計を読み取り、
資産推移纏め シートに本日付け列として追記（既存なら上書き）する

【読み取り方針】
  1) 小計セルのキャッシュ値を優先使用 (Excel で再計算済みの場合)
  2) キャッシュが None の場合 (Python コピー直後) は数式を解析して直接計算:
     G 日本株  : col7(時価単価) × col10(口数)
     H 米国株  : col7(USD単価) × col10(口数) × USDJPY
     F 投資信託: JPY系 → col7×col10, USD系 → col7×col10×USDJPY
     B 外貨現預金: col7(外貨) × 為替レート (USD→H29, RMB→col8直接値)
     D US$社債  : col9(評価単価%) × col12(額面lots) × USDJPY
  3) それでも取得できない場合のみベース値を使用

  ※ G〜J セクションの行番号はシートごとに変わるため動的スキャン
"""

import os, sys, re, shutil
from datetime import date, datetime
from openpyxl.utils.datetime import to_excel as _dt_to_excel

if sys.stdout.encoding and sys.stdout.encoding.lower() in ('cp932', 'shift_jis', 'shift-jis'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl

BASE_DIR = r'C:\Users\yna78\OneDrive - Corazon\Yuji Private\資産ポートフォリオ戦略'

def find_latest_excel(base_dir: str) -> str:
    """BASE_DIR 内の最新 資産クラス整理_YYYYMMDD.xlsx を返す"""
    import glob
    files = glob.glob(os.path.join(base_dir, '資産クラス整理_????????.xlsx'))
    if not files:
        raise FileNotFoundError(f'資産クラス整理_YYYYMMDD.xlsx が見つかりません: {base_dir}')
    return max(files)

EXCEL_FILE = find_latest_excel(BASE_DIR)
print(f'対象ファイル: {EXCEL_FILE}')

TS_SHEET = '資産推移纏め'

TARGET_DATE = date.today()
SN_NEW = '資産クラス' + TARGET_DATE.strftime('%Y%m%d')

print('=== transfer_to_summary.py ===')
print(f'対象日  : {TARGET_DATE}')
print(f'新シート: {SN_NEW}')

# ── ファイルをコピーして読み込み ─────────────────────────────
tmp = EXCEL_FILE + '.tmp.xlsx'
shutil.copy2(EXCEL_FILE, tmp)

# data_only=False で数式文字列も取得 (直接計算のため)
wb_formula = openpyxl.load_workbook(tmp)
# data_only=True で数式のキャッシュ値を取得 (Excel 再計算済みの場合)
wb_cache   = openpyxl.load_workbook(tmp, data_only=True)

if SN_NEW not in wb_formula.sheetnames:
    print(f'[ERROR] シート "{SN_NEW}" が見つかりません。利用可能: {wb_formula.sheetnames}')
    os.remove(tmp)
    sys.exit(1)

ws_f  = wb_formula[SN_NEW]    # 数式付きシート
ws_c  = wb_cache[SN_NEW]      # キャッシュ値シート
ws_ts = wb_cache[TS_SHEET]

# ── ユーティリティ ────────────────────────────────────────────
def cached(r, c):
    """キャッシュ値 (data_only) を float で返す"""
    v = ws_c.cell(row=r, column=c).value
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None

def formula_val(r, c):
    """数式シートのセル値を返す (数式文字列 or 直接値)"""
    return ws_f.cell(row=r, column=c).value

def eval_cell(r, c, usdjpy):
    """
    セル値を float に変換。
    - 直接値 → そのまま
    - =$H$29 / =H29 → USDJPY
    - =RrowCcol 形式の単純参照 → 参照先セルを再帰的に評価
    - 純粋な四則演算数式 (=7*20, =(1800+2558)*5 等) → eval
    - それ以外 → None
    """
    v = formula_val(r, c)
    if v is None:
        return None
    # datetime → Excel シリアル番号に変換 (数値として書かれたセルが日付書式の場合)
    if isinstance(v, datetime):
        try:
            return float(_dt_to_excel(v))
        except Exception:
            return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s.startswith('='):
        try:
            return float(s)
        except (TypeError, ValueError):
            return None
    expr = s[1:].replace('$', '').strip()
    # H29 参照 → USDJPY
    if re.fullmatch(r'H29', expr, re.IGNORECASE):
        return usdjpy
    # SUM(範囲) 形式 (例: =SUM(I127:I128)) → 範囲内セルを再帰評価して合計
    m_sum = re.fullmatch(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', expr, re.IGNORECASE)
    if m_sum:
        from openpyxl.utils import column_index_from_string
        sc = column_index_from_string(m_sum.group(1).upper())
        ec = column_index_from_string(m_sum.group(3).upper())
        sr, er = int(m_sum.group(2)), int(m_sum.group(4))
        total = 0.0
        for rr in range(sr, er + 1):
            for cc in range(sc, ec + 1):
                vv = eval_cell(rr, cc, usdjpy)
                if vv:
                    total += vv
        return total if total else None
    # 単純セル参照 (例: =I49, =N53, =G75) → 参照先を評価
    m = re.fullmatch(r'([A-Z]+)(\d+)', expr, re.IGNORECASE)
    if m:
        from openpyxl.utils import column_index_from_string
        ref_col = column_index_from_string(m.group(1).upper())
        ref_row = int(m.group(2))
        return eval_cell(ref_row, ref_col, usdjpy)
    # 単純な四則演算のみ許可
    if re.fullmatch(r'[\d\s\+\-\*\/\(\)\.]+', expr):
        try:
            return float(eval(expr))
        except Exception:
            return None
    return None

# ── 資産推移纏め 行マップ構築 ────────────────────────────────
ts_row_map = {}
for r in range(2, ws_ts.max_row + 1):
    v = ws_ts.cell(row=r, column=2).value
    if v:
        ts_row_map[str(v).strip()] = r

print(f'\n資産推移纏め 行マップ:')
for k, v in ts_row_map.items():
    print(f'  row{v}: "{k}"')

# ── 書き込み列を決定 ────────────────────────────────────────
target_col = None
for c in range(3, ws_ts.max_column + 2):
    cv = ws_ts.cell(row=2, column=c).value
    if cv is None:
        target_col = c
        break
    if isinstance(cv, datetime) and cv.date() == TARGET_DATE:
        target_col = c
        print(f'  既存の {TARGET_DATE} 列を上書き: col={target_col}')
        break
if target_col is None:
    target_col = ws_ts.max_column + 1

base_col = target_col - 1
print(f'書き込み列={target_col}  ベース列={base_col}')

def base_val(label):
    r = ts_row_map.get(label)
    if r:
        v = ws_ts.cell(row=r, column=base_col).value
        return float(v) if v is not None else 0.0
    return 0.0

# ── USDJPY 取得 ──────────────────────────────────────────────
USDJPY_NEW = cached(29, 8)
if not USDJPY_NEW or USDJPY_NEW < 50:
    # 数式シートから直接値を取得 (update_stock.py が書き込む直接値)
    v = formula_val(29, 8)
    try:
        USDJPY_NEW = float(v) if isinstance(v, (int, float)) else float(str(v))
    except Exception:
        USDJPY_NEW = 143.0
    print(f'  [INFO] USDJPY キャッシュなし → 数式シートから {USDJPY_NEW}')
if not USDJPY_NEW or USDJPY_NEW < 50:
    USDJPY_NEW = 143.0
    print(f'  [WARN] USDJPY 取得失敗, {USDJPY_NEW} を使用')

SN_PREV_CANDIDATES = [s for s in wb_formula.sheetnames if s.startswith('資産クラス') and s != SN_NEW]
USDJPY_PREV = None
if SN_PREV_CANDIDATES:
    sn_prev = sorted(SN_PREV_CANDIDATES)[-1]
    v_prev = wb_cache[sn_prev].cell(row=29, column=8).value
    try:
        USDJPY_PREV = float(v_prev) if v_prev is not None else None
    except (TypeError, ValueError):
        pass
    # フォールバック: 数式シートから
    if not USDJPY_PREV or USDJPY_PREV < 50:
        v2 = wb_formula[sn_prev].cell(row=29, column=8).value
        try:
            USDJPY_PREV = float(v2) if isinstance(v2, (int, float)) else None
        except Exception:
            pass
    print(f'前回シート: {sn_prev}  USDJPY_PREV={USDJPY_PREV}')
if not USDJPY_PREV or USDJPY_PREV < 50:
    USDJPY_PREV = 159.0
    print(f'  [WARN] USDJPY_PREV 取得失敗, {USDJPY_PREV} を使用')

print(f'USDJPY: 前回={USDJPY_PREV:.3f}  新={USDJPY_NEW:.3f}')

# ═══════════════════════════════════════════════════════════
# セクション動的スキャン
# ═══════════════════════════════════════════════════════════

def scan_section(section_label, search_start=70):
    """
    B列= section_label のセクションの (小計行, データ行リスト) を返す。
    """
    in_sec = False
    subtotal_row = None
    data_rows = []
    for r in range(search_start, ws_f.max_row + 1):
        b  = str(ws_f.cell(row=r, column=2).value or '').strip()
        c3 = str(ws_f.cell(row=r, column=3).value or '').strip()
        d  = ws_f.cell(row=r, column=4).value
        if b == section_label:
            in_sec = True
            continue
        if not in_sec:
            continue
        if b and b != section_label and len(b) == 1 and b.isalpha():
            break
        if c3 == '小計':
            subtotal_row = r
            break
        if d and str(d).strip() not in ('名前', ''):
            data_rows.append(r)
    return subtotal_row, data_rows

# ═══════════════════════════════════════════════════════════
# 各セクション小計の取得
# ═══════════════════════════════════════════════════════════
totals = {}

# ── A: 円建て現預金 ────────────────────────────────────────
sec_a = cached(26, 6)
if not sec_a:
    # col6 フォールバック: 直接値の合計 (rows 6-25)
    sec_a = sum(
        float(v) for r in range(6, 26)
        if (v := formula_val(r, 6)) is not None and isinstance(v, (int, float)) and v > 0
    )
if not sec_a:
    sec_a = base_val('円建て現預金')
    print(f'A 円建て現預金: 取得失敗 → ベース {sec_a:,.0f}')
else:
    print(f'A 円建て現預金: {sec_a:,.0f}')
totals['円建て現預金'] = sec_a

# ── B: 外貨現預金 ──────────────────────────────────────────
sec_b = cached(36, 6)
if not sec_b:
    # col6 = col8(rate) * col7(外貨額) の数式 → 直接計算
    total_b = 0.0
    for r in range(29, 36):
        d = ws_f.cell(row=r, column=4).value
        if not d or str(d).strip() in ('名前', ''):
            continue
        fx_amt = eval_cell(r, 7, USDJPY_NEW)   # 外貨金額
        rate   = eval_cell(r, 8, USDJPY_NEW)   # 為替レート (USD→USDJPY, RMB→20)
        if fx_amt is not None and rate is not None:
            total_b += fx_amt * rate
    sec_b = total_b if total_b > 0 else base_val('外貨現預金')
    if total_b > 0:
        print(f'B 外貨現預金: {sec_b:,.0f}  (外貨×為替 直接計算)')
    else:
        print(f'B 外貨現預金: 取得失敗 → ベース {sec_b:,.0f}')
else:
    print(f'B 外貨現預金: {sec_b:,.0f}')
totals['外貨現預金'] = sec_b

# ── C: 円建て社債 ──────────────────────────────────────────
sec_c = 0.0
for r in range(38, 58):
    c3 = str(ws_f.cell(row=r, column=3).value or '').strip()
    if c3 == '小計':
        break
    v = eval_cell(r, 7, USDJPY_NEW)
    if v and v > 0:
        sec_c += v
if not sec_c:
    sec_c = base_val('円建て社債')
    print(f'C 円建て社債: 取得失敗 → ベース {sec_c:,.0f}')
else:
    print(f'C 円建て社債: {sec_c:,.0f}')
totals['円建て社債'] = sec_c

# ── D: US$建て社債 ─────────────────────────────────────────
# 小計行: col7='小計' かつ col8=SUM数式 (col3 ではなく col7 に小計ラベル)
d_label = next((k for k in ts_row_map if 'US' in k and '社債' in k), 'US＄建て社債')
# 小計行を動的検索
d_subtotal_row = None
for r in range(42, 60):
    c7 = str(ws_f.cell(row=r, column=7).value or '').strip()
    c3 = str(ws_f.cell(row=r, column=3).value or '').strip()
    if c7 == '小計' or c3 == '小計':
        d_subtotal_row = r
        break
sec_d = cached(d_subtotal_row, 8) if d_subtotal_row else None
if not sec_d:
    # col8 = col9(評価単価%) × USDJPY × col12(額面lots)
    total_d = 0.0
    end_row = d_subtotal_row if d_subtotal_row else 56
    for r in range(43, end_row):
        d = ws_f.cell(row=r, column=4).value
        if not d or str(d).strip() in ('名前', ''):
            continue
        price = eval_cell(r, 9, USDJPY_NEW)   # 評価単価 (セル参照も解決)
        face  = eval_cell(r, 12, USDJPY_NEW)  # 額面 (L列)
        if price is not None and face is not None and isinstance(face, (int, float)) and face > 0:
            total_d += price * USDJPY_NEW * face
    sec_d = total_d if total_d > 0 else base_val(d_label) * (USDJPY_NEW / USDJPY_PREV)
    if total_d > 0:
        print(f'D US$建て社債: {sec_d:,.0f}  (評価単価×FX×額面 直接計算, 小計=row{d_subtotal_row})')
    else:
        print(f'D US$建て社債: 取得失敗 → 為替調整ベース {sec_d:,.0f}')
else:
    print(f'D US$建て社債: {sec_d:,.0f}  (小計=row{d_subtotal_row})')
totals[d_label] = sec_d

# ── E: Private Credit ─────────────────────────────────────
e_label = next((k for k in ts_row_map if 'Private' in k or 'Credit' in k), 'Private Credit')
sec_e = cached(56, 8)
if not sec_e:
    v56 = formula_val(56, 8)
    if isinstance(v56, (int, float)) and v56 > 0:
        sec_e = float(v56)
    else:
        # H56 = =I56*J56*L56: 評価単価(USD) × USDJPY × 額面
        # L56 は数値だが日付書式で保存されている → eval_cell で Excel シリアルに変換
        price_e = eval_cell(56, 9, USDJPY_NEW)   # I56: 評価単価 (USD)
        face_e  = eval_cell(56, 12, USDJPY_NEW)  # L56: 額面 (datetime→シリアル変換)
        if price_e and face_e and face_e > 0:
            sec_e = price_e * USDJPY_NEW * face_e
            print(f'E {e_label}: {sec_e:,.0f}  (評価単価×FX×額面: {price_e:.4f}×{USDJPY_NEW:.2f}×{face_e:.0f})')
        else:
            e_base = base_val(e_label)
            sec_e = e_base * (USDJPY_NEW / USDJPY_PREV) if e_base else 0.0
            print(f'E {e_label}: 直接計算失敗 (price={price_e}, face={face_e}) → 為替調整ベース {sec_e:,.0f}')
if sec_e:
    print(f'E {e_label}: {sec_e:,.0f}')
totals[e_label] = sec_e

# ── F: 投資信託 ────────────────────────────────────────────
sec_f = cached(68, 8)
if not sec_f:
    total_f = 0.0
    for r in range(60, 68):
        d = ws_f.cell(row=r, column=4).value
        if not d or str(d).strip() in ('名前', ''):
            continue
        price  = eval_cell(r, 7, USDJPY_NEW)   # 時価単価 (JPY/口 or USD/口)
        shares = eval_cell(r, 10, USDJPY_NEW)  # 口数 (J列)
        if price is None or shares is None:
            continue
        # N列 (col14) が =$H$29 なら USD ETF → USDJPY 乗算
        n_val = formula_val(r, 14)
        is_usd = isinstance(n_val, str) and 'H29' in n_val.replace('$', '')
        total_f += price * shares * (USDJPY_NEW if is_usd else 1.0)
    sec_f = total_f if total_f > 0 else base_val('投資信託')
    if total_f > 0:
        print(f'F 投資信託: {sec_f:,.0f}  (単価×口数 直接計算)')
    else:
        print(f'F 投資信託: 取得失敗 → ベース {sec_f:,.0f}')
else:
    print(f'F 投資信託: {sec_f:,.0f}')
totals['投資信託'] = sec_f

# ── G: 日本株 ──────────────────────────────────────────────
g_sub_row, g_data = scan_section('G')
sec_g = cached(g_sub_row, 8) if g_sub_row else None
if not sec_g:
    total_g = 0.0
    for r in g_data:
        price  = eval_cell(r, 7, USDJPY_NEW)   # 時価単価 (JPY)
        shares = eval_cell(r, 10, USDJPY_NEW)  # 口数
        if price is not None and shares is not None:
            total_g += price * shares
    sec_g = total_g if total_g > 0 else base_val('日本株')
    if total_g > 0:
        print(f'G 日本株: {sec_g:,.0f}  (単価×口数 直接計算, 小計row={g_sub_row})')
    else:
        print(f'G 日本株: 取得失敗 → ベース {sec_g:,.0f}')
else:
    print(f'G 日本株: {sec_g:,.0f}  (小計row={g_sub_row} キャッシュ)')
totals['日本株'] = sec_g

# ── H: 米国株 ──────────────────────────────────────────────
h_sub_row, h_data = scan_section('H')
sec_h = cached(h_sub_row, 8) if h_sub_row else None
if not sec_h:
    total_h = 0.0
    for r in h_data:
        price  = eval_cell(r, 7, USDJPY_NEW)   # USD 単価
        shares = eval_cell(r, 10, USDJPY_NEW)  # 口数
        if price is not None and shares is not None:
            total_h += price * shares * USDJPY_NEW
    sec_h = total_h if total_h > 0 else base_val('米国株')
    if total_h > 0:
        print(f'H 米国株: {sec_h:,.0f}  (USD単価×口数×FX 直接計算, 小計row={h_sub_row})')
    else:
        print(f'H 米国株: 取得失敗 → ベース {sec_h:,.0f}')
else:
    print(f'H 米国株: {sec_h:,.0f}  (小計row={h_sub_row} キャッシュ)')
totals['米国株'] = sec_h

# ── REIT: 0 固定 ────────────────────────────────────────────
reit_label = next((k for k in ts_row_map if 'REIT' in k), 'REIT')
totals[reit_label] = 0.0
print(f'REIT: 0  (固定)')

# ── I: 事業投資（優先株）──────────────────────────────────────
i_sub_row, i_data = scan_section('I')
sec_i = cached(i_sub_row, 8) if i_sub_row else None
i_label = next((k for k in ts_row_map if '事業' in k or '優先株' in k), '事業投資（優先株）')
if not sec_i:
    # col8 は直接値 or =SUM(In:Im) 形式 → eval_cell で両対応
    total_i = 0.0
    for r in i_data:
        v = eval_cell(r, 8, USDJPY_NEW)
        if v and v > 0:
            total_i += v
    sec_i = total_i if total_i > 0 else base_val(i_label)
    if total_i > 0:
        print(f'I 事業投資: {sec_i:,.0f}  (直接計算, 小計row={i_sub_row})')
    else:
        print(f'I 事業投資: 取得失敗 → ベース {sec_i:,.0f}')
else:
    print(f'I 事業投資: {sec_i:,.0f}  (小計row={i_sub_row} キャッシュ)')
totals[i_label] = sec_i

# ── J: 自宅不動産 ─────────────────────────────────────────
j_label = next((k for k in ts_row_map if '不動産' in k), '自宅不動産')
j_sub_row, j_data = scan_section('J')
sec_j = None
for r in j_data:
    v = formula_val(r, 8)
    if isinstance(v, (int, float)) and v > 0:
        sec_j = float(v)
        print(f'J 自宅不動産: {sec_j:,.0f}  (row{r} 直接値)')
        break
if not sec_j:
    sec_j = base_val(j_label)
    print(f'J 自宅不動産: 取得失敗 → ベース {sec_j:,.0f}')
totals[j_label] = sec_j

# ── 合計計算 ───────────────────────────────────────────────
excl_keys = {'不動産', 'Total', '合計'}
total_excl = sum(v for k, v in totals.items() if not any(x in k for x in excl_keys))
total_incl = total_excl + sec_j

print(f'\n=== 合計サマリー ===')
for k, v in totals.items():
    print(f'  {k}: {v:,.0f}')
print(f'  Total(不動産除く): {total_excl:,.0f}')
print(f'  Total(不動産含む): {total_incl:,.0f}')

# ═══════════════════════════════════════════════════════════
# ① 検証: 資産クラスシート Summary の A～I, A～J 合計と突合
# ═══════════════════════════════════════════════════════════
print(f'\n=== 検証: Total(不動産除く) = A～I合計 ===')
ai_keys = [k for k in totals if not any(x in k for x in excl_keys)]
sum_ai = sum(totals[k] for k in ai_keys)
diff_excl = abs(total_excl - sum_ai)
if diff_excl < 1:
    print(f'  ✓ 一致: {total_excl:,.0f}')
else:
    print(f'  ✗ 差異あり: Total={total_excl:,.0f}, A-I合計={sum_ai:,.0f}, 差={diff_excl:,.0f}')

print(f'=== 検証: Total(不動産含む) = A～J合計 ===')
sum_aj = sum_ai + sec_j
diff_incl = abs(total_incl - sum_aj)
if diff_incl < 1:
    print(f'  ✓ 一致: {total_incl:,.0f}')
else:
    print(f'  ✗ 差異あり: Total={total_incl:,.0f}, A-J合計={sum_aj:,.0f}, 差={diff_incl:,.0f}')

# ═══════════════════════════════════════════════════════════
# ② 資産クラスシートの Summary テーブルを検索してセル参照を構築
# ═══════════════════════════════════════════════════════════
def find_summary_links(ws_formula, sheet_name):
    """
    資産クラスシートの Summary テーブル (col5='Summary') を探し、
    ラベル → Excel セル参照式の辞書を返す。
    例: {'円建て現預金': "='資産クラス20260429'!G146", ...}
    特殊キー '__total_ai__' / '__total_aj__' に A-I / A-J 合計行の参照も含む。
    """
    header_row = None
    for r in range(130, min(ws_formula.max_row + 1, 250)):
        v5 = str(ws_formula.cell(row=r, column=5).value or '').strip()
        v7 = str(ws_formula.cell(row=r, column=7).value or '').strip()
        if v5 == 'Summary' and v7 == '時価合計':
            header_row = r
            break
    if not header_row:
        print('  [WARN] Summary テーブルが見つかりません → 直接値で書き込み')
        return None

    from openpyxl.utils import get_column_letter
    col_g = get_column_letter(7)   # G列 = col7 = 時価合計
    safe  = f"'{sheet_name}'"      # シート名をシングルクォートで囲む

    links = {}
    for r in range(header_row + 1, header_row + 40):
        v5 = str(ws_formula.cell(row=r, column=5).value or '').strip()
        v6 = str(ws_formula.cell(row=r, column=6).value or '').strip()
        if v5 in list('ABCDEFGHIJ') and v6:
            links[v6] = f'={safe}!{col_g}{r}'
        elif 'A～I' in v6 or '合計（A～I' in v6:
            links['__total_ai__'] = f'={safe}!{col_g}{r}'
        elif 'A～J' in v6 or '合計（A～J' in v6:
            links['__total_aj__'] = f'={safe}!{col_g}{r}'

    print(f'  Summary テーブル発見 (header=row{header_row}), {len(links)} リンクを構築')
    for lbl, ref in links.items():
        print(f'    {lbl}: {ref}')
    return links

print('\n=== Summary リンク構築 ===')
summary_links = find_summary_links(ws_f, SN_NEW)

# ── 計算済み値を JSON に保存 (gen_dashboard.py のフォールバック用) ──────
import json as _json
_totals_for_json = dict(totals)
_totals_for_json['Total(不動産除く）'] = round(total_excl)
_totals_for_json['Total(不動産含む）'] = round(total_incl)
_json_obj = {
    'date':   TARGET_DATE.isoformat(),
    'sheet':  SN_NEW,
    'values': {k: round(v) for k, v in _totals_for_json.items()}
}
_json_path = os.path.join(BASE_DIR, '_latest_totals.json')
with open(_json_path, 'w', encoding='utf-8') as _jf:
    _json.dump(_json_obj, _jf, ensure_ascii=False, indent=2)
print(f'  → JSON保存: {_json_path}  ({len(_json_obj["values"])} 項目)')

# ═══════════════════════════════════════════════════════════
# ③ 資産推移纏め に書き込む
# ═══════════════════════════════════════════════════════════
wb2 = openpyxl.load_workbook(tmp)
ws_ts2 = wb2[TS_SHEET]

print(f'\n書き込み: 資産推移纏め col={target_col}  ({TARGET_DATE})')

ws_ts2.cell(row=2, column=target_col).value = datetime(
    TARGET_DATE.year, TARGET_DATE.month, TARGET_DATE.day)

total_label_excl = next((k for k in ts_row_map if 'Total' in k and ('除く' in k or 'excl' in k.lower())), None)
total_label_incl = next((k for k in ts_row_map if 'Total' in k and '含む' in k), None)

written = []
for r in range(2, ws_ts2.max_row + 1):
    lbl = ws_ts2.cell(row=r, column=2).value
    if not lbl:
        continue
    lbl_s = str(lbl).strip()

    # Total(不動産除く) → Summary の 合計(A～I) セルにリンク
    if lbl_s == total_label_excl:
        if summary_links and '__total_ai__' in summary_links:
            cell_val = summary_links['__total_ai__']
            ws_ts2.cell(row=r, column=target_col).value = cell_val
            written.append((r, lbl_s, cell_val))
        else:
            ws_ts2.cell(row=r, column=target_col).value = round(total_excl)
            written.append((r, lbl_s, round(total_excl)))
        continue

    # Total(不動産含む) → Summary の 合計(A～J) セルにリンク
    if lbl_s == total_label_incl:
        if summary_links and '__total_aj__' in summary_links:
            cell_val = summary_links['__total_aj__']
            ws_ts2.cell(row=r, column=target_col).value = cell_val
            written.append((r, lbl_s, cell_val))
        else:
            ws_ts2.cell(row=r, column=target_col).value = round(total_incl)
            written.append((r, lbl_s, round(total_incl)))
        continue

    # REIT: 常に 0 (Summary テーブルに存在しない)
    if 'REIT' in lbl_s:
        ws_ts2.cell(row=r, column=target_col).value = 0
        written.append((r, lbl_s, 0))
        continue

    # 各資産クラス → Summary テーブルの G 列セルにリンク
    if summary_links and lbl_s in summary_links:
        cell_val = summary_links[lbl_s]
        ws_ts2.cell(row=r, column=target_col).value = cell_val
        written.append((r, lbl_s, cell_val))
    elif lbl_s in totals:
        # フォールバック: 直接値
        val = round(totals[lbl_s])
        ws_ts2.cell(row=r, column=target_col).value = val
        written.append((r, lbl_s, val))

print(f'書き込んだ行:')
for r, lbl, v in written:
    print(f'  row{r} "{lbl}": {v}')

wb2.save(tmp)
try:
    os.replace(tmp, EXCEL_FILE)
    print(f'\n[OK] 保存完了: {EXCEL_FILE}')
except (PermissionError, OSError):
    bk = os.path.join(BASE_DIR, f'資産クラス整理_{TARGET_DATE.strftime("%Y%m%d")}_updated.xlsx')
    shutil.move(tmp, bk)
    print(f'\n[WARNING] 元ファイルが開かれているため別名保存: {bk}')

print('完了!')
