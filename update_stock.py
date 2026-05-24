# -*- coding: utf-8 -*-
"""
update_stock.py
毎週土曜日に実行: 株価・為替を取得してExcelを更新する

データソース:
  ひふみ投信  : Yahoo Finance Japan  https://finance.yahoo.co.jp/quote/9C31108A
  VTI/VOO/QQQ : Bloomberg Japan (403 の場合 yfinance にフォールバック)
  その他      : yfinance
"""

import os
import sys
import shutil
import re
from datetime import date

# Windows cp932 コンソール対策
if sys.stdout.encoding and sys.stdout.encoding.lower() in ('cp932', 'shift_jis', 'shift-jis'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import requests
from bs4 import BeautifulSoup
import yfinance as yf
import openpyxl

BASE_DIR   = r'C:\Users\yna78\OneDrive - Corazon\Yuji Private\資産ポートフォリオ戦略'
EXCEL_FILE = os.path.join(BASE_DIR, '資産クラス整理_20260524.xlsx')

def find_latest_sheet(wb) -> str:
    """直近の「資産クラスYYYYMMDD」シートを返す（今日のシートは除く）"""
    today_str = date.today().strftime('%Y%m%d')
    candidates = sorted(
        [s for s in wb.sheetnames
         if s.startswith('資産クラス') and s[5:].isdigit()
         and len(s[5:]) == 8 and s[5:] != today_str],
        reverse=True
    )
    if not candidates:
        raise RuntimeError('コピー元となる資産クラスシートが見つかりません')
    return candidates[0]

SCRAPE_HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/124.0.0.0 Safari/537.36'
    ),
    'Accept-Language': 'ja-JP,ja;q=0.9,en-US;q=0.8',
}

# ──────────────────────────────────────────────
# yfinance tickers → [(row, col), ...]
#   col 7 = G (時価単価)
#   col 8 = H (直近為替 H29)
#
# ※ 資産クラス整理_20260524.xlsx (repair_bak ベース) の行構造:
#   G セクション(日本株): row72〜row82
#     ・東京地下鉄(9023)が row75/76 の 2行、IHI(7013) が row77/78 の 2行
#       → 両行とも同一銘柄コードで直接値を書き込む
#   H セクション(米国株): row87 からデータ開始 (row86=ヘッダ)
#     ・row93,97,99,106,115,121 等は数式参照行 (=G92 等) のため更新不要
# ──────────────────────────────────────────────
TICKER_CELLS = {
    # 直近為替
    'USDJPY=X': [(29, 8)],
    # 投資信託 ETF (Bloomberg 優先 → yfinance フォールバック、別関数で処理)
    'VTI': [(62, 7)],
    'VOO': [(63, 7)],
    'QQQ': [(64, 7)],
    # 日本株 (G セクション: row72〜82)
    '8001.T': [(72, 7)],          # 伊藤忠商事
    '8316.T': [(73, 7)],          # 三井住友FG
    '8306.T': [(74, 7)],          # 三菱UFJ
    '9023.T': [(75, 7), (76, 7)], # 東京地下鉄(9023) × 2行
    '7013.T': [(77, 7), (78, 7)], # IHI(7013) × 2行
    '8303.T': [(79, 7)],          # SBI新生銀行
    '2501.T': [(80, 7)],          # サッポロHD
    '285A.T': [(81, 7)],          # キオクシア
    '7011.T': [(82, 7)],          # 三菱重工
    # 米国株 (H セクション: row87 からデータ開始、数式参照行は除外)
    'GOOG':  [(87, 7)],           # row99/100-102/116/121 は =G87 で自動反映
    'ISRG':  [(88, 7)],
    'AMZN':  [(89, 7)],           # row103/104 は =G89 で自動反映
    'MSFT':  [(90, 7)],           # row97/110/117/118 は =G90 で自動反映
    'NVDA':  [(91, 7)],           # row106/108 は =G91 で自動反映
    'META':  [(92, 7)],           # row93/107/109 は =G92 で自動反映
    'GS':    [(94, 7)],           # Goldman Sachs; row119 は =G94 で自動反映
    'JPM':   [(95, 7)],           # JP Morgan
    'LLY':   [(96, 7)],
    'PLTR':  [(98, 7)],
    'AAPL':  [(105, 7)],          # シートでは APPL と誤記; row112/115 は =G105 で自動反映
    'TSLA':  [(113, 7)],
    'EW':    [(114, 7)],
    'MS':    [(120, 7)],          # Morgan Stanley
    # PAWN (row122) は上場廃止のため除外
}

BLOOMBERG_ETFS = {
    'VTI': 'https://www.bloomberg.com/jp/quote/VTI:US',
    'VOO': 'https://www.bloomberg.com/jp/quote/VOO:US',
    'QQQ': 'https://www.bloomberg.com/jp/quote/QQQ:US',
}


def get_price_yfinance(ticker: str) -> float | None:
    try:
        hist = yf.Ticker(ticker).history(period='5d')
        if not hist.empty:
            return round(float(hist['Close'].iloc[-1]), 4)
    except Exception as e:
        print(f'  [WARN] yfinance {ticker}: {e}')
    return None


def get_hifumi_price() -> float | None:
    """Yahoo Finance Japan からひふみ投信の基準価額を取得 (yen/1万口 → yen/口)"""
    url = 'https://finance.yahoo.co.jp/quote/9C31108A'
    try:
        r = requests.get(url, headers=SCRAPE_HEADERS, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, 'html.parser')
        el = soup.select_one(
            'span.StyledNumber__value__3rXW, '
            'span[class*="PriceBoard__price"], '
            'span[class*="StyledNumber--vertical"]'
        )
        if el:
            raw = el.get_text().strip().replace(',', '')
            nav_per_10k = float(raw)          # 円/1万口
            return round(nav_per_10k / 10000, 4)   # 円/1口
    except Exception as e:
        print(f'  [WARN] ひふみ投信スクレイプ失敗: {e}')
    return None


def get_bloomberg_etf_price(ticker: str) -> float | None:
    """Bloomberg Japan からETF価格を取得。失敗時は yfinance にフォールバック"""
    url = BLOOMBERG_ETFS[ticker]
    try:
        r = requests.get(url, headers=SCRAPE_HEADERS, timeout=15)
        if r.status_code == 403:
            raise ValueError('Bloomberg 403 - フォールバック')
        r.raise_for_status()
        soup = BeautifulSoup(r.text, 'html.parser')
        # Bloomberg price selectors (layout は変わる可能性あり)
        for sel in ['[class*="priceText"]', '[class*="price__"]', '[data-field="regularMarketPrice"]']:
            el = soup.select_one(sel)
            if el:
                raw = re.sub(r'[^\d.]', '', el.get_text())
                if raw:
                    return round(float(raw), 4)
    except Exception as e:
        print(f'  [INFO] Bloomberg {ticker} ({url}): {e} → yfinance にフォールバック')
    return get_price_yfinance(ticker)


def copy_worksheet(wb, source_name: str, dest_name: str):
    if dest_name in wb.sheetnames:
        del wb[dest_name]
    ws = wb.copy_worksheet(wb[source_name])
    ws.title = dest_name
    return ws


def main():
    today        = date.today()
    new_sheet    = f'資産クラス{today.strftime("%Y%m%d")}'
    update_label = f'更新日: {today.strftime("%Y/%m/%d")}'

    print(f'=== update_stock.py 実行日: {today} ===')
    print(f'ソースファイル: {EXCEL_FILE}')

    if not os.path.exists(EXCEL_FILE):
        print(f'[ERROR] ファイルが見つかりません: {EXCEL_FILE}')
        sys.exit(1)

    tmp_path = EXCEL_FILE + '.tmp.xlsx'
    shutil.copy2(EXCEL_FILE, tmp_path)

    wb = openpyxl.load_workbook(tmp_path)

    # コピー元: 直近の日付シートを自動選択
    source_sheet = find_latest_sheet(wb)
    print(f'ソースシート  : {source_sheet}  →  {new_sheet}')

    print(f'\nシートをコピー: {source_sheet} → {new_sheet}')
    ws = copy_worksheet(wb, source_sheet, new_sheet)

    ws.cell(row=1, column=1).value = update_label
    print(f'A1 に "{update_label}" を記入')

    print('\n--- 価格取得・更新 ---')

    # ひふみ投信 (Yahoo Finance Japan)
    hifumi_price = get_hifumi_price()
    if hifumi_price is not None:
        ws.cell(row=60, column=7).value = hifumi_price
        print(f'  ひふみ投信 (Yahoo Finance Japan): {hifumi_price} 円/口')
    else:
        print('  ひふみ投信: 取得失敗 → スキップ (要手動更新)')

    # VTI / VOO / QQQ (Bloomberg → yfinance フォールバック)
    for etf in ['VTI', 'VOO', 'QQQ']:
        price = get_bloomberg_etf_price(etf)
        if price is not None:
            for (row, col) in TICKER_CELLS[etf]:
                ws.cell(row=row, column=col).value = price
            print(f'  {etf}: {price}')
        else:
            print(f'  {etf}: 取得失敗 → スキップ')

    # yfinance (残りのティッカー)
    skip_etfs = {'VTI', 'VOO', 'QQQ'}
    for ticker, cells in TICKER_CELLS.items():
        if ticker in skip_etfs:
            continue
        price = get_price_yfinance(ticker)
        if price is None:
            print(f'  {ticker}: データ取得失敗 → スキップ')
            continue
        for (row, col) in cells:
            ws.cell(row=row, column=col).value = price
        label = 'USDJPY' if ticker == 'USDJPY=X' else ticker
        print(f'  {label}: {price}')

    # AB米国成長株投信: yfinance 非対応
    print('\n  [要手動更新] AB米国成長株投信(Hedgeなし) (G65): yfinance/Bloomberg 非対応')

    # 保存
    wb.save(tmp_path)

    try:
        if os.path.exists(EXCEL_FILE):
            os.replace(tmp_path, EXCEL_FILE)
        else:
            shutil.move(tmp_path, EXCEL_FILE)
        print(f'\n[OK] 保存完了: {EXCEL_FILE}')
    except (PermissionError, OSError):
        backup = os.path.join(BASE_DIR, f'資産整理_{today.strftime("%Y%m%d")}.xlsx')
        shutil.move(tmp_path, backup)
        print(f'\n[WARNING] 元ファイルが開かれているため別名で保存: {backup}')
    print(f'  新しいシート: {new_sheet}')


if __name__ == '__main__':
    main()
