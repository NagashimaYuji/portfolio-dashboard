# -*- coding: utf-8 -*-
"""
gen_dashboard.py
ポートフォリオダッシュボード HTML を生成する
"""

import os, sys, json, re
from datetime import datetime

if sys.stdout.encoding and sys.stdout.encoding.lower() in ('cp932', 'shift_jis', 'shift-jis'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.utils import get_column_letter

BASE_DIR   = r'C:\Users\yna78\OneDrive - Corazon\Yuji Private\資産ポートフォリオ戦略'
EXCEL_FILE = os.path.join(BASE_DIR, '資産クラス整理_20260524.xlsx')
OUT_HTML   = os.path.join(BASE_DIR, 'dashboard.html')

TS_SHEET  = '資産推移纏め'

print('Excelを読み込み中...')
wb         = openpyxl.load_workbook(EXCEL_FILE, data_only=True)   # キャッシュ値
wb_formula = openpyxl.load_workbook(EXCEL_FILE)                   # 数式文字列

# ──────────────────────────────────────────────────────────
# クロスシート数式リゾルバ
# 例: ='資産クラス20260429'!G146 → wb[sheet].cell(row, col).value
# 最大 3 段階追跡 (G146=F26=SUM(...) のような参照チェーンに対応)
# ──────────────────────────────────────────────────────────
import re as _re
from openpyxl.utils import column_index_from_string as _col_idx

def resolve_cell(sheet_name, row, col, max_depth=4):
    """
    指定シートのセル値を解決する。キャッシュ優先 → 数式チェーンを再帰追跡。
    sheet_name: 対象シート名 (wb/wb_formula の両方に存在すること)
    """
    if max_depth <= 0:
        return None
    # data_only キャッシュを優先
    if sheet_name in wb.sheetnames:
        v = wb[sheet_name].cell(row=row, column=col).value
        if v is not None and not str(v).startswith('='):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None
    # 数式文字列を取得して再帰
    if sheet_name not in wb_formula.sheetnames:
        return None
    formula = wb_formula[sheet_name].cell(row=row, column=col).value
    if formula is None:
        return None
    s = str(formula).strip().replace('$', '')
    if not s.startswith('='):
        try:
            return float(s)
        except (TypeError, ValueError):
            return None
    expr = s[1:]
    # クロスシート参照: 'SheetName'!GN または SheetName!GN
    m_x = _re.match(r"'?([^'!]+)'?!([A-Z]+)(\d+)", expr, _re.IGNORECASE)
    if m_x:
        return resolve_cell(m_x.group(1), int(m_x.group(3)),
                            _col_idx(m_x.group(2).upper()), max_depth - 1)
    # 同シート内の単純参照: G146 や F26
    m_s = _re.fullmatch(r'([A-Z]+)(\d+)', expr, _re.IGNORECASE)
    if m_s:
        return resolve_cell(sheet_name, int(m_s.group(2)),
                            _col_idx(m_s.group(1).upper()), max_depth - 1)
    # 数値算術式: 7*20, 10*20, 100+50 など (セル参照なし)
    if _re.fullmatch(r'[\d\s\+\-\*\/\(\)\.]+', expr):
        try:
            return float(eval(expr))  # noqa: S307 – 数字と演算子のみ許可
        except Exception:
            pass
    return None

def resolve_formula_value(formula_str, _unused_ws=None, max_depth=4):
    """
    資産推移纏め の数式文字列 (='SheetName'!GN) を解決して数値を返す。
    """
    if formula_str is None:
        return None
    s = str(formula_str).strip().replace('$', '')
    if not s.startswith('='):
        try:
            return float(s)
        except (TypeError, ValueError):
            return None
    expr = s[1:]
    m = _re.match(r"'?([^'!]+)'?!([A-Z]+)(\d+)", expr, _re.IGNORECASE)
    if m:
        return resolve_cell(m.group(1), int(m.group(3)),
                            _col_idx(m.group(2).upper()), max_depth)
    return None

# 最新の「資産クラスYYYYMMDD」シートを自動選択
_sn_candidates = sorted(
    [s for s in wb.sheetnames if s.startswith('資産クラス') and s[5:].isdigit()
     and len(s[5:]) == 8],
    reverse=True
)
SN_SHEET = _sn_candidates[0] if _sn_candidates else '資産クラス20260415'
print(f'  スナップショットシート: {SN_SHEET}')

# ══════════════════════════════════════════════════════════
# 1. 資産推移纏め  (time series)
#    row2 col2='資産クラス', col3以降=datetime
#    row3以降 col2=ラベル, col3以降=値
# ══════════════════════════════════════════════════════════
ws_ts = wb[TS_SHEET]
MAX_COL = ws_ts.max_column

# 日付行 (row 2, col 3 ～ MAX_COL)
dates = []
date_cols = []
for c in range(3, MAX_COL+1):
    v = ws_ts.cell(row=2, column=c).value
    if v is None:
        continue
    if isinstance(v, datetime):
        dates.append(v.strftime('%Y-%m-%d'))
        date_cols.append(c)
    elif isinstance(v, (int, float)):
        # 年ラベルっぽい数値は無視
        pass

print(f'  日付数: {len(dates)}, 最初={dates[0] if dates else "?"}, 最後={dates[-1] if dates else "?"}')

# シリーズ行 (row 3 ～ max_row)
ts_series = {}
ts_labels_order = []
for r in range(3, ws_ts.max_row+1):
    lbl = ws_ts.cell(row=r, column=2).value
    if not lbl:
        continue
    lbl = str(lbl).strip()
    vals = []
    for c in date_cols:
        v = ws_ts.cell(row=r, column=c).value
        try:
            vals.append(round(float(v)/1e6, 3) if v is not None else None)
        except (TypeError, ValueError):
            vals.append(None)
    ts_series[lbl] = vals
    ts_labels_order.append(lbl)

print(f'  シリーズ: {ts_labels_order}')

# ══════════════════════════════════════════════════════════
# 2. スナップショット (資産クラス20260415)
# ══════════════════════════════════════════════════════════
ws_sn = wb[SN_SHEET]

def sv(r, c):
    v = ws_sn.cell(row=r, column=c).value
    return str(v).strip() if v is not None else ''

def fv(r, c):
    v = ws_sn.cell(row=r, column=c).value
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None

usdjpy    = fv(29, 8) or 143.0
upd_date  = sv(1, 1)
print(f'  USDJPY={usdjpy}, 更新日={upd_date}')

# ──────────────────────────────────────────────────────────
# 2-a. セクション境界を検出 (B列 = 'A'〜'J' のとき新セクション)
# ──────────────────────────────────────────────────────────
# セクション定義: {key: {label, header_row, data_rows:[row_idx,...]}}
sections = {}
sec_order = []
current = None

for r in range(1, ws_sn.max_row+1):
    b = sv(r, 2)
    c3 = sv(r, 3)

    if re.match(r'^[A-J]$', b):
        current = b
        if current not in sections:
            sections[current] = {'label': c3, 'header_row': None, 'data_rows': []}
            sec_order.append(current)
        # セクションヘッダー行自体にデータがある場合 (例: E行にPrivate Creditデータ)
        d_hdr = ws_sn.cell(row=r, column=4).value
        if d_hdr is not None and str(d_hdr).strip() not in ('名前', ''):
            sections[current]['data_rows'].append(r)
    elif current:
        # 小計行 (C列が '小計') → セクション終端
        if c3 == '小計':
            current = None
            continue
        # 空行をスキップ
        # D列に値がある行を data_row とみなす
        d = ws_sn.cell(row=r, column=4).value
        if d is not None:
            row_type = sv(r, 4)   # D列
            # D列が '名前' のような見出し行をスキップ
            if row_type in ('名前',):
                sections[current]['header_row'] = r
            else:
                sections[current]['data_rows'].append(r)

print(f'  セクション: {[(k, sections[k]["label"], len(sections[k]["data_rows"])) for k in sec_order]}')

# ──────────────────────────────────────────────────────────
# 2-b. 各セクションのデータを抽出
# ──────────────────────────────────────────────────────────
# セクションA (円建て現預金): D=名前, E=金融機関, F=金額(円)
# セクションB (外貨現預金):  D=名前, E=金融機関, F=通貨, G=金額(外貨), H=レート, I=円換算
# セクションC (円建て社債):  D=名前, E=金融機関, F=銘柄, G=時価単価, H=時価総額, I=購入単価, J=口数, K=購入総額
# セクションD (US$建て社債): 同上
# セクションE (Private Credit): 同上
# セクションF (投資信託): 同上
# セクションG (日本株):    D=名前, E=金融機関, F=銘柄, G=時価単価, H=時価総額, I=購入単価, J=口数, K=購入総額
# セクションH (米国株):    同上
# セクションI (事業投資):  個別カラム違う可能性あり
# セクションJ (自宅不動産): 個別

def cell_num(r, c, follow_ref=True):
    """data_only → formula直接値 → セル参照解決 の順で数値を取得"""
    v = wb[SN_SHEET].cell(row=r, column=c).value
    if isinstance(v, (int, float)):
        return float(v)
    v2 = wb_formula[SN_SHEET].cell(row=r, column=c).value
    if isinstance(v2, (int, float)):
        return float(v2)
    # 数式参照を追跡 (=G72, =I49 等)
    if follow_ref and isinstance(v2, str) and v2.startswith('='):
        resolved = resolve_cell(SN_SHEET, r, c, max_depth=3)
        if resolved is not None:
            return resolved
    return None

def extract_sec_a(rows):
    """円建て現預金: D=名前, E=金融機関, F=金額"""
    out = []
    for r in rows:
        name  = sv(r, 4)
        inst  = sv(r, 5)
        amt   = fv(r, 6)
        out.append({'row':r,'name':name,'institution':inst,'amount':amt,
                    'quantity':None,'cost_unit':None,'cost_total':None,'price':None,'value':amt})
    return out

def extract_sec_b(rows):
    """外貨現預金:
    col4=名前, col5=金融機関, col9=通貨テキスト, col7=外貨金額, col8=為替レート(formula chain)
    col6 is JPY formula (=H*G) → None in data_only
    """
    out = []
    for r in rows:
        name   = sv(r, 4)
        inst   = sv(r, 5)
        ccy    = sv(r, 9)   # col9 = 通貨テキスト ('US$', 'RMB')
        fx_amt = fv(r, 7)   # col7 = 外貨金額(直接値)
        # col8 = 為替レート: =H29 or =H30→H29 等 (最大4段階チェーン) → depth=6 で解決
        rate   = cell_num(r, 8)
        if rate is None:
            rate = resolve_cell(SN_SHEET, r, 8, max_depth=6)
        if rate is None and ccy == 'US$':
            rate = usdjpy  # フォールバック
        jpy    = round(fx_amt * rate) if fx_amt and rate else None
        out.append({'row':r,'name':name,'institution':inst,'currency':ccy,
                    'fx_amount':fx_amt,'rate':rate,'jpy_amount':jpy,
                    'quantity':fx_amt,'cost_unit':rate,'cost_total':jpy,'price':rate,'value':jpy})
    return out

def extract_sec_c(rows):
    """円建て社債:
    col4=名前, col5=金融機関, col6=銘柄, col7=評価総額(直接値), col11=購入総額(直接値)
    """
    out = []
    for r in rows:
        name       = sv(r, 4)
        inst       = sv(r, 5)
        brand      = sv(r, 6)
        eval_total = fv(r, 7)    # col7 = 評価総額(直接値)
        purch_total= fv(r, 11)   # col11 = 購入総額(直接値)
        gain = round(eval_total - purch_total) if (eval_total is not None and purch_total is not None) else None
        out.append({'row':r,'name':name,'institution':inst,'brand':brand,
                    'eval_total':eval_total,'purch_total':purch_total,'gain':gain,
                    'price':None,'value':eval_total,'cost_unit':None,'cost_total':purch_total,'quantity':None})
    return out

def extract_sec_d(rows):
    """US$建て社債:
    col4=名前, col5=金融機関, col6=銘柄
    col9=評価単価%(直接値), col10=直近為替(formula=H29), col12=額面数量(直接値)
    col13=購入総額(直接値 or formula), col15=購入時為替(直接値)
    """
    out = []
    for r in rows:
        name      = sv(r, 4)
        inst      = sv(r, 5)
        brand     = sv(r, 6)
        eval_pct  = cell_num(r, 9)          # 評価単価%
        fx        = cell_num(r, 10) or usdjpy  # 直近為替 (formula=H29 解決)
        face      = cell_num(r, 12)          # 額面数量
        purch_total = cell_num(r, 13)        # 購入総額 (直接値 or formula)
        purch_fx  = fv(r, 15)               # col15 = 購入時為替 (直接値)
        # col13 が formula の場合: =N*O*L = col14(購入単価%) × col15(購入時為替) × col12(額面)
        if purch_total is None:
            p_pct = fv(r, 14)   # col14 = 購入単価%
            if p_pct and purch_fx and face:
                purch_total = round(p_pct * purch_fx * face)
        eval_total = round(eval_pct * fx * face) if (eval_pct and fx and face) else None
        gain = round(eval_total - purch_total) if (eval_total is not None and purch_total is not None) else None
        out.append({'row':r,'name':name,'institution':inst,'brand':brand,
                    'eval_price':eval_pct,'fx':fx,'face':face,
                    'eval_total':eval_total,'purch_total':purch_total,'gain':gain,'purch_fx':purch_fx,
                    'price':eval_pct,'value':eval_total,'cost_unit':None,'cost_total':purch_total,'quantity':face})
    return out

def extract_sec_generic(rows):
    """汎用 (E/I/J): D=名前, E=金融機関, F=銘柄, G=時価単価, H=時価総額, I=購入単価, J=口数, K=購入総額"""
    out = []
    for r in rows:
        name   = sv(r, 4)
        inst   = sv(r, 5)
        brand  = sv(r, 6)
        price  = fv(r, 7)
        value  = fv(r, 8)
        c_unit = fv(r, 9)
        qty    = fv(r, 10)
        c_tot  = fv(r, 11)
        out.append({'row':r,'name':name,'institution':inst,'brand':brand,
                    'quantity':qty,'cost_unit':c_unit,'cost_total':c_tot,'price':price,'value':value})
    return out

def extract_sec_f_full(rows):
    """投資信託:
    col7=評価単価(直接値 or formula), col9=購入単価, col10=数量, col14=直近為替formula($H$29=USD)
    col8=評価総額(formula=G*J or G*J*N → None), col11=購入合計(formula → None)
    """
    out = []
    for r in rows:
        name       = sv(r, 4)
        inst       = sv(r, 5)
        brand      = sv(r, 6)
        eval_price = cell_num(r, 7)          # 評価単価 (follow formula ref)
        purch_price= fv(r, 9) or cell_num(r, 9)  # 購入単価
        qty        = fv(r, 10) or cell_num(r, 10)  # 数量
        # col14 に =$H$29 が入っていればUSD建て
        n_fml = wb_formula[SN_SHEET].cell(row=r, column=14).value
        is_usd = isinstance(n_fml, str) and 'H29' in n_fml.replace('$', '')
        fx = usdjpy if is_usd else None
        # 評価総額: キャッシュ優先、なければ計算
        eval_total = cell_num(r, 8)
        if eval_total is None and eval_price and qty:
            eval_total = round(eval_price * qty * (usdjpy if is_usd else 1.0))
        # 購入合計: キャッシュ優先、なければ計算
        purch_fx_val = fv(r, 15) or usdjpy if is_usd else None
        purch_total = cell_num(r, 11)
        if purch_total is None and purch_price and qty:
            purch_total = round(purch_price * qty * (purch_fx_val if is_usd else 1.0))
        gain = round(eval_total - purch_total) if (eval_total is not None and purch_total is not None) else None
        out.append({'row':r,'name':name,'institution':inst,'brand':brand,
                    'eval_price':eval_price,'fx':fx,'quantity':qty,
                    'eval_total':eval_total,'purch_total':purch_total,'gain':gain,
                    'price':eval_price,'value':eval_total,'cost_unit':purch_price,'cost_total':purch_total})
    return out

def extract_sec_g(rows):
    """日本株:
    col7=時価単価(直接値), col8=時価総額(formula=G*J → None)
    col9=購入単価(直接値 or formula), col10=口数(直接値 or formula), col11=購入総額(formula → None)
    """
    out = []
    for r in rows:
        name   = sv(r, 4)
        inst   = sv(r, 5)
        brand  = sv(r, 6)
        price  = fv(r, 7) or cell_num(r, 7)
        c_unit = cell_num(r, 9)   # 購入単価 (formula解決)
        qty    = cell_num(r, 10)  # 口数 (formula解決: =7*20 等)
        value  = cell_num(r, 8)
        if value is None and price and qty:
            value = round(price * qty)
        c_tot = cell_num(r, 11)
        if c_tot is None and c_unit and qty:
            c_tot = round(c_unit * qty)
        gain = round(value - c_tot) if (value is not None and c_tot is not None) else None
        out.append({'row':r,'name':name,'institution':inst,'brand':brand,
                    'price':price,'quantity':qty,'value':value,
                    'cost_unit':c_unit,'cost_total':c_tot,'gain':gain})
    return out

def extract_sec_h(rows):
    """米国株:
    col7=USD単価(直接値 or formula=Gxx), col8=時価総額(formula=G*J*N → None)
    col9=購入単価(直接値 or formula), col10=口数(直接値 or formula)
    col11=購入総額(formula → None), col15=購入時為替(直接値)
    """
    out = []
    for r in rows:
        name    = sv(r, 4)
        inst    = sv(r, 5)
        brand   = sv(r, 6)
        price   = fv(r, 7) or cell_num(r, 7)   # USD単価 (formula解決)
        c_unit  = cell_num(r, 9)                # 購入単価
        qty     = cell_num(r, 10)               # 口数 (formula解決)
        purch_fx= fv(r, 15)                     # 購入時為替
        value   = cell_num(r, 8)
        if value is None and price and qty:
            value = round(price * qty * usdjpy)
        c_tot = cell_num(r, 11)
        if c_tot is None and c_unit and qty:
            c_tot = round(c_unit * qty * (purch_fx or usdjpy))
        gain = round(value - c_tot) if (value is not None and c_tot is not None) else None
        out.append({'row':r,'name':name,'institution':inst,'brand':brand,
                    'price':price,'quantity':qty,'value':value,
                    'cost_unit':c_unit,'cost_total':c_tot,'gain':gain,'purch_fx':purch_fx})
    return out

# セクションFは投資信託 (AB米国成長株 = cols 9,10)
# Let's check section F more carefully
print('\nセクションF行確認:')
if 'F' in sections:
    for r in sections['F']['data_rows'][:5]:
        row_d = [(c, repr(ws_sn.cell(row=r,column=c).value)[:20]) for c in range(4,13) if ws_sn.cell(row=r,column=c).value is not None]
        print(f'  row{r}: {row_d}')

# セクション別抽出
detail = {}
for sk in sec_order:
    rows = sections[sk]['data_rows']
    lbl  = sections[sk]['label']
    if sk == 'A':
        rows_data = extract_sec_a(rows)
    elif sk == 'B':
        rows_data = extract_sec_b(rows)
    elif sk == 'C':
        rows_data = extract_sec_c(rows)
    elif sk == 'D':
        rows_data = extract_sec_d(rows)
    elif sk == 'E':
        rows_data = extract_sec_generic(rows)
    elif sk == 'F':
        rows_data = extract_sec_f_full(rows)
    elif sk == 'G':
        rows_data = extract_sec_g(rows)
    elif sk == 'H':
        rows_data = extract_sec_h(rows)
    else:
        rows_data = extract_sec_generic(rows)
    detail[sk] = {'label': lbl, 'rows': rows_data}

# ──────────────────────────────────────────────────────────
# 2-c. 人別合計
# ──────────────────────────────────────────────────────────
# H列(時価総額)は数式セルでキャッシュがないため、価格×口数×FXから直接計算する
# ──────────────────────────────────────────────────────────
PERSON_MAP = {
    'yuji': 'Yuji', 'ゆうじ': 'Yuji',
    'fumi': 'Fumie', 'fumie': 'Fumie', 'ふみえ': 'Fumie',
    'mina': 'Mina', 'みな': 'Mina',
    'akatsuki': 'Akatsuki', 'akasuki': 'Akatsuki', 'あかつき': 'Akatsuki',  # akasuki = typo in Excel
}

def get_person(name_str):
    nl = name_str.lower()
    for k, v in PERSON_MAP.items():
        if k in nl:
            return v
    return None

person_totals = {'Yuji': 0.0, 'Fumie': 0.0, 'Mina': 0.0, 'Akatsuki': 0.0}

# A: 円建て現預金 → col6(F)=残高(円) 直接値
for rd in detail.get('A', {}).get('rows', []):
    p = get_person(rd.get('name',''))
    v = rd.get('value')
    if p and v:
        person_totals[p] += v

# B: 外貨現預金 → fx_amount(col7) × rate (col8, formula chain解決済み)
for rd in detail.get('B', {}).get('rows', []):
    p = get_person(rd.get('name',''))
    if p:
        fx   = rd.get('fx_amount') or cell_num(rd['row'], 7)
        rate = rd.get('rate') or cell_num(rd['row'], 8) or usdjpy
        if fx and rate:
            person_totals[p] += fx * rate

# C: 円建て社債 → eval_total (col7=評価総額、直接値)
for rd in detail.get('C', {}).get('rows', []):
    p = get_person(rd.get('name',''))
    if p:
        v = rd.get('eval_total') or cell_num(rd['row'], 7)
        if v:
            person_totals[p] += v

# D: US$建て社債 → extract_sec_d で計算済みの eval_total を使用
# ※ フォールバックとして評価単価(col9) × USDJPY × 額面(col12) も試みる
from openpyxl.utils.datetime import to_excel as _dte
from datetime import datetime as _dt
def _face_val(r):
    v = cell_num(r, 12)
    if v is None:
        raw = wb_formula[SN_SHEET].cell(row=r, column=12).value
        if isinstance(raw, _dt):
            try: v = float(_dte(raw))
            except Exception: pass
    return v

for rd in detail.get('D', {}).get('rows', []):
    p = get_person(rd.get('name',''))
    if p:
        v = rd.get('eval_total')
        if v is None:
            price_pct = cell_num(rd['row'], 9)
            face      = _face_val(rd['row'])
            if price_pct and face:
                v = price_pct * usdjpy * face
        if v:
            person_totals[p] += v

# E: Private Credit → col9(I=評価単価) × USDJPY × col12(L=額面)
for rd in detail.get('E', {}).get('rows', []):
    p = get_person(rd.get('name',''))
    if p:
        price_pct = rd.get('cost_unit') or cell_num(rd['row'], 9)
        face      = _face_val(rd['row'])
        if price_pct and face:
            person_totals[p] += price_pct * usdjpy * face

# F: 投資信託 → extract_sec_f_full で計算済みの eval_total を使用
for rd in detail.get('F', {}).get('rows', []):
    p = get_person(rd.get('name',''))
    if p:
        v = rd.get('eval_total')
        if v is None:
            price = cell_num(rd['row'], 7) or cell_num(rd['row'], 9)
            qty   = rd.get('quantity') or cell_num(rd['row'], 10)
            if price and qty:
                n_fml = wb_formula[SN_SHEET].cell(row=rd['row'], column=14).value
                is_usd = n_fml and 'H29' in str(n_fml).replace('$', '')
                v = price * qty * (usdjpy if is_usd else 1.0)
        if v:
            person_totals[p] += v

# G: 日本株 → extract_sec_g で計算済みの value を使用
for rd in detail.get('G', {}).get('rows', []):
    p = get_person(rd.get('name',''))
    if p:
        v = rd.get('value')
        if v is None:
            price = rd.get('price') or cell_num(rd['row'], 7)
            qty   = rd.get('quantity') or cell_num(rd['row'], 10)
            if price and qty:
                v = price * qty
        if v:
            person_totals[p] += v

# H: 米国株 → extract_sec_h で計算済みの value を使用
for rd in detail.get('H', {}).get('rows', []):
    p = get_person(rd.get('name',''))
    if p:
        v = rd.get('value')
        if v is None:
            price = rd.get('price') or cell_num(rd['row'], 7)
            qty   = rd.get('quantity') or cell_num(rd['row'], 10)
            if price and qty:
                v = price * qty * usdjpy
        if v:
            person_totals[p] += v

# I: 事業投資 → col8(H) キャッシュ or col7(G=単価)×col10(J=口数)
for rd in detail.get('I', {}).get('rows', []):
    p = get_person(rd.get('name',''))
    if p:
        v = rd.get('value') or cell_num(rd['row'], 8)
        if v is None:
            price = cell_num(rd['row'], 7)
            qty   = cell_num(rd['row'], 10)
            if price and qty:
                v = price * qty
        if v:
            person_totals[p] += v

pers_sum = sum(person_totals.values())
print(f'  人別合計: {person_totals}')
print(f'  人別合計計: {pers_sum:,.0f}')

# ──────────────────────────────────────────────────────────
# 2-d. 資産クラス別合計 → 資産推移纏め の最新データ列から
# ──────────────────────────────────────────────────────────
# 最後に値が入っている列を探す (None が続く前の列)
LAST_COL = max(date_cols) if date_cols else MAX_COL
print(f'  最新データ列: {LAST_COL}  ({dates[-1] if dates else "?"})')

class_totals_raw = {}
ws_ts_f = wb_formula[TS_SHEET]   # 数式文字列シート (resolve 用)

# JSON フォールバック: transfer_to_summary.py が出力した計算済み値
import json as _json, os as _os
_json_fallback = {}
_json_path = _os.path.join(BASE_DIR, '_latest_totals.json')
if _os.path.exists(_json_path):
    try:
        with open(_json_path, encoding='utf-8') as _jf:
            _jd = _json.load(_jf)
        if _jd.get('sheet') == SN_SHEET:
            _json_fallback = {k: float(v) for k, v in _jd.get('values', {}).items()}
            print(f'  JSONフォールバック読込: {_jd.get("date")} ({len(_json_fallback)}項目)')
    except Exception as _e:
        print(f'  [WARN] JSON読込失敗: {_e}')

for r in range(3, ws_ts.max_row+1):
    lbl = ws_ts.cell(row=r, column=2).value
    val = ws_ts.cell(row=r, column=LAST_COL).value
    if not lbl:
        continue
    if val is None:
        # Step1: 数式チェーンを最大4段階追跡して解決
        formula_str = ws_ts_f.cell(row=r, column=LAST_COL).value
        val = resolve_formula_value(formula_str, max_depth=4)
    if val is None and _json_fallback:
        # Step2: JSON フォールバック (formula キャッシュが全 None の場合)
        val = _json_fallback.get(str(lbl).strip())
    if val is not None:
        try:
            class_totals_raw[str(lbl).strip()] = float(val)
        except (TypeError, ValueError):
            pass

print(f'  資産クラス別合計 (raw): {class_totals_raw}')

# 百万円に変換
class_totals = {k: round(v/1e6, 2) for k, v in class_totals_raw.items()}

# ══════════════════════════════════════════════════════════
# 3. USD vs JPY 分類
# ══════════════════════════════════════════════════════════
USD_KEYWORDS = ['米国株', 'US$', 'US＄', 'USD']  # 全角＄も含む
SKIP_KEYWORDS = ['合計', 'Total', '不動産', 'REIT']  # 流動資産計算から除外

usd_total = 0.0
jpy_total = 0.0
for lbl, val in class_totals.items():
    if any(kw in lbl for kw in SKIP_KEYWORDS):
        continue  # 合計行・非流動資産は除外
    if any(kw in lbl for kw in USD_KEYWORDS):
        usd_total += val
    else:
        jpy_total += val

# 流動資産 = Total(不動産除く) (資産推移纏めの値を優先、なければ usd+jpy)
liquid_total = next(
    (v for k, v in class_totals.items() if 'Total' in k and '除く' in k),
    round(usd_total + jpy_total, 2)
)
print(f'  USD建て: {usd_total:.1f}M, 円建て: {jpy_total:.1f}M, 流動資産: {liquid_total:.1f}M')

# ══════════════════════════════════════════════════════════
# 4. JSON データ保存
# ══════════════════════════════════════════════════════════
# detail の rows を JSON シリアライズ可能に変換
def clean_rows(rows):
    out = []
    for rd in rows:
        cleaned = {}
        for k, v in rd.items():
            if isinstance(v, float) and (v != v):  # NaN check
                cleaned[k] = None
            else:
                cleaned[k] = v
        out.append(cleaned)
    return out

detail_json = {}
for sk, sv_data in detail.items():
    detail_json[sk] = {'label': sv_data['label'], 'rows': clean_rows(sv_data['rows'])}

dashboard_data = {
    'update_date': upd_date,
    'usdjpy': usdjpy,
    'ts_dates': dates,
    'ts_series': ts_series,
    'ts_labels_order': ts_labels_order,
    'class_totals': class_totals,
    'person_totals': {k: round(v/1e6, 2) for k, v in person_totals.items()},
    'usd_total': round(usd_total, 2),
    'jpy_total': round(jpy_total, 2),
    'detail': detail_json,
}

json_path = os.path.join(BASE_DIR, '_dashboard_data.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(dashboard_data, f, ensure_ascii=False, indent=2, default=str)
print(f'\nデータ保存: {json_path}')

# ══════════════════════════════════════════════════════════
# 5. HTML 生成
# ══════════════════════════════════════════════════════════
print('HTML生成中...')

PALETTE = [
    '#4e79a7','#f28e2b','#e15759','#76b7b2','#59a14f',
    '#edc948','#b07aa1','#ff9da7','#9c755f','#bab0ac',
    '#17becf','#aec7e8','#ffbe7d',
]

# Chart.js datasets
def mk_datasets(series_keys, series_dict):
    ds = []
    for i, lbl in enumerate(series_keys):
        color = PALETTE[i % len(PALETTE)]
        vals  = series_dict.get(lbl, [])
        ds.append({
            'label': lbl,
            'data': vals,
            'borderColor': color,
            'backgroundColor': color + '55',
            'fill': True,
            'tension': 0.3,
            'pointRadius': 0,
            'borderWidth': 1.5,
        })
    return ds

# 合計系を除いたシリーズだけ表示 (グラフ見やすく)
DISPLAY_SERIES = [l for l in ts_labels_order if '合計' not in l]
TOTAL_SERIES   = [l for l in ts_labels_order if '合計' in l]

datasets_main  = mk_datasets(DISPLAY_SERIES, ts_series)
datasets_total = mk_datasets(TOTAL_SERIES,   ts_series)

# 評価損益の色付けヘルパー
def gain_cls(v):
    if v is None:
        return ''
    return ' text-success fw-semibold' if v > 0 else ' text-danger fw-semibold'

# 詳細タブ HTML
def mk_sec_tabs():
    nav = ''
    content = ''
    for idx, sk in enumerate(sec_order):
        if sk not in detail_json:
            continue
        sec = detail_json[sk]
        lbl = sec['label']
        tab_id = f'sec{sk}'
        active = ' active' if idx == 0 else ''
        show   = ' show active' if idx == 0 else ''

        # ヘッダー列の決定
        if sk == 'A':
            heads = ['名前', '金融機関', '金額(円)']
        elif sk == 'B':
            heads = ['名前', '金融機関', '通貨', '外貨額', '為替レート', '円換算']
        elif sk == 'C':
            heads = ['名前', '金融機関', '銘柄', '評価総額', '購入合計', '評価損益']
        elif sk == 'D':
            heads = ['名前', '金融機関', '銘柄', '評価単価(%)', '直近為替', '額面', '評価総額', '購入合計', '評価損益']
        elif sk == 'F':
            heads = ['名前', '金融機関', '銘柄', '評価単価', '直近為替', '数量', '評価総額', '購入合計', '評価損益']
        elif sk == 'G':
            heads = ['名前', '金融機関', '銘柄', '時価単価', '口数', '時価総額', '購入単価', '購入総額', '評価損益']
        elif sk == 'H':
            heads = ['名前', '金融機関', '銘柄', '時価単価(USD)', '直近為替', '口数', '時価総額', '購入単価(USD)', '購入総額', '評価損益']
        else:
            # E/I/J generic
            heads = ['名前', '金融機関', '銘柄', '時価単価', '時価総額', '購入単価', '口数', '購入総額']

        thead = ''.join(f'<th>{h}</th>' for h in heads)
        tbody = ''
        for rd in sec['rows']:
            # row_html はクロージャ問題を避けるためインライン展開
            if sk == 'A':
                cells = (f"<td>{rd.get('name','')}</td><td>{rd.get('institution','')}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('value'))}</td>")
            elif sk == 'B':
                cells = (f"<td>{rd.get('name','')}</td><td>{rd.get('institution','')}</td>"
                         f"<td>{rd.get('currency','')}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('fx_amount'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('rate'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('jpy_amount'))}</td>")
            elif sk == 'C':
                g = rd.get('gain')
                cells = (f"<td>{rd.get('name','')}</td><td>{rd.get('institution','')}</td>"
                         f"<td>{rd.get('brand','')}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('eval_total'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('purch_total'))}</td>"
                         f"<td class='text-end{gain_cls(g)}'>{fmt_num(g)}</td>")
            elif sk == 'D':
                g = rd.get('gain')
                cells = (f"<td>{rd.get('name','')}</td><td>{rd.get('institution','')}</td>"
                         f"<td>{rd.get('brand','')}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('eval_price'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('fx'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('face'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('eval_total'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('purch_total'))}</td>"
                         f"<td class='text-end{gain_cls(g)}'>{fmt_num(g)}</td>")
            elif sk == 'F':
                g = rd.get('gain')
                fx_disp = fmt_num(rd.get('fx')) if rd.get('fx') else '&#x2015;'
                cells = (f"<td>{rd.get('name','')}</td><td>{rd.get('institution','')}</td>"
                         f"<td>{rd.get('brand','')}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('eval_price'))}</td>"
                         f"<td class='text-end'>{fx_disp}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('quantity'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('eval_total'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('purch_total'))}</td>"
                         f"<td class='text-end{gain_cls(g)}'>{fmt_num(g)}</td>")
            elif sk == 'G':
                g = rd.get('gain')
                cells = (f"<td>{rd.get('name','')}</td><td>{rd.get('institution','')}</td>"
                         f"<td>{rd.get('brand','')}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('price'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('quantity'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('value'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('cost_unit'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('cost_total'))}</td>"
                         f"<td class='text-end{gain_cls(g)}'>{fmt_num(g)}</td>")
            elif sk == 'H':
                g = rd.get('gain')
                cells = (f"<td>{rd.get('name','')}</td><td>{rd.get('institution','')}</td>"
                         f"<td>{rd.get('brand','')}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('price'))}</td>"
                         f"<td class='text-end'>{fmt_num(usdjpy)}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('quantity'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('value'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('cost_unit'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('cost_total'))}</td>"
                         f"<td class='text-end{gain_cls(g)}'>{fmt_num(g)}</td>")
            else:
                # E/I/J generic
                cells = (f"<td>{rd.get('name','')}</td><td>{rd.get('institution','')}</td>"
                         f"<td>{rd.get('brand','')}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('price'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('value'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('cost_unit'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('quantity'))}</td>"
                         f"<td class='text-end'>{fmt_num(rd.get('cost_total'))}</td>")

            tbody += f"<tr data-row='{rd['row']}'>{cells}<td><button class='btn btn-xs btn-outline-primary py-0 px-1 edit-btn' onclick='startEdit(this)'>編集</button><button class='btn btn-xs btn-success py-0 px-1 save-btn d-none' onclick='saveEdit(this)'>保存</button><button class='btn btn-xs btn-secondary py-0 px-1 cancel-btn d-none' onclick='cancelEdit(this)'>取消</button></td></tr>"

        nav += f'<li class="nav-item"><button class="nav-link{active} py-1 px-2" id="{tab_id}-tab" data-bs-toggle="tab" data-bs-target="#{tab_id}" type="button">{sk}: {lbl}</button></li>'
        content += f'''<div class="tab-pane fade{show}" id="{tab_id}" role="tabpanel">
          <div class="d-flex justify-content-between align-items-center my-2">
            <h6 class="mb-0 fw-bold">{sk}: {lbl}</h6>
            <button class="btn btn-sm btn-outline-secondary py-0" onclick="exportSection('{sk}')">CSV</button>
          </div>
          <div class="table-responsive">
            <table class="table table-hover table-sm table-bordered small" id="table-{sk}">
              <thead class="table-dark"><tr>{thead}<th>操作</th></tr></thead>
              <tbody>{tbody}</tbody>
            </table>
          </div></div>'''

    return nav, content

def fmt_num(v):
    if v is None:
        return ''
    try:
        f = float(v)
        if f == int(f) and abs(f) < 1e12:
            return f'{int(f):,}'
        return f'{f:,.2f}'
    except (TypeError, ValueError):
        return str(v)

nav_html, content_html = mk_sec_tabs()

# JSON埋め込み用
dates_j   = json.dumps(dates, ensure_ascii=False)
ds_main_j = json.dumps(datasets_main, ensure_ascii=False)
ds_tot_j  = json.dumps(datasets_total, ensure_ascii=False)

# サマリー用 (Total行・合計行はclsから除外してtotに入れる)
cls_lbl_j = json.dumps([k for k in class_totals.keys()
                         if '合計' not in k and 'Total' not in k], ensure_ascii=False)
cls_val_j = json.dumps([v for k,v in class_totals.items()
                         if '合計' not in k and 'Total' not in k], ensure_ascii=False)
tot_lbl_j = json.dumps([k for k in class_totals.keys()
                         if 'Total' in k or '合計' in k], ensure_ascii=False)
tot_val_j = json.dumps([v for k,v in class_totals.items()
                         if 'Total' in k or '合計' in k], ensure_ascii=False)

pers_lbl_j = json.dumps(list(dashboard_data['person_totals'].keys()), ensure_ascii=False)
pers_val_j = json.dumps(list(dashboard_data['person_totals'].values()), ensure_ascii=False)

usd_jpy_lbl = json.dumps(['USD建て資産(円換算)', '円建て資産'], ensure_ascii=False)
usd_jpy_val = json.dumps([dashboard_data['usd_total'], dashboard_data['jpy_total']], ensure_ascii=False)
liquid_total_j = json.dumps(round(liquid_total, 2))
# 総資産（不動産含む）= Total(不動産含む)
total_incl = next(
    (v for k, v in class_totals.items() if 'Total' in k and '含む' in k),
    round(liquid_total + class_totals.get('自宅不動産', 0.0), 2)
)
total_incl_j = json.dumps(round(total_incl, 2))

html = f'''<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>資産ポートフォリオ ダッシュボード</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-adapter-date-fns@3.0.0/dist/chartjs-adapter-date-fns.bundle.min.js"></script>
<style>
body{{font-family:"Meiryo","Yu Gothic",sans-serif;background:#f0f2f5;font-size:.88rem;}}
.navbar-brand{{font-size:1.1rem;}}
.chart-tall{{height:400px;}}
.chart-md{{height:280px;}}
.kpi-card{{background:#fff;border-radius:10px;padding:1rem 1.2rem;box-shadow:0 2px 6px rgba(0,0,0,.08);text-align:center;}}
.kpi-val{{font-size:1.6rem;font-weight:700;color:#0d6efd;}}
.kpi-lbl{{font-size:.78rem;color:#666;margin-top:.2rem;}}
.sec-header{{background:linear-gradient(135deg,#0d6efd,#6610f2);color:#fff;padding:.4rem .9rem;border-radius:6px;margin-bottom:.8rem;font-weight:600;}}
.editable.editing{{background:#fffde7;}}
.editable.editing input{{width:100%;border:1px solid #0d6efd;border-radius:3px;padding:1px 4px;}}
#toast-box{{position:fixed;bottom:18px;right:18px;z-index:9999;}}
.btn-xs{{font-size:.72rem;line-height:1.2;}}
.table-sm td,.table-sm th{{padding:.25rem .4rem;}}
</style>
</head>
<body>



<nav class="navbar navbar-dark bg-dark px-3 py-2">
  <span class="navbar-brand fw-bold">📊 資産ポートフォリオ ダッシュボード</span>
  <span class="text-light small">{upd_date} &nbsp;|&nbsp; USD/JPY: {usdjpy:.2f}</span>
</nav>

<div class="container-fluid px-3 py-3">
<ul class="nav nav-tabs mb-3 fw-semibold" id="mainTab">
  <li class="nav-item"><button class="nav-link active" data-bs-toggle="tab" data-bs-target="#tab-ts">📈 資産推移纏め</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tab-sum">📋 サマリー</button></li>
  <li class="nav-item"><button class="nav-link" data-bs-toggle="tab" data-bs-target="#tab-det">🗂 資産クラス詳細</button></li>
</ul>

<div class="tab-content">

<!-- ① 資産推移纏め -->
<div class="tab-pane fade show active" id="tab-ts">
  <div class="bg-white rounded shadow-sm p-3 mb-3">
    <div class="sec-header">📈 資産推移 (百万円) — クラス別積み上げ</div>
    <div class="d-flex gap-1 mb-2 flex-wrap">
      <button class="btn btn-sm btn-outline-secondary" onclick="setRange('all')">全期間</button>
      <button class="btn btn-sm btn-outline-secondary" onclick="setRange('5y')">5年</button>
      <button class="btn btn-sm btn-outline-secondary" onclick="setRange('3y')">3年</button>
      <button class="btn btn-sm btn-outline-secondary" onclick="setRange('1y')">1年</button>
      <button class="btn btn-sm btn-outline-secondary" onclick="setRange('6m')">6ヶ月</button>
      <button class="btn btn-sm btn-outline-secondary" onclick="setRange('1m')">1ヶ月</button>
      <div class="ms-3 d-flex align-items-center gap-2">
        <label class="small mb-0">合計線:</label>
        <div class="form-check form-switch mb-0">
          <input class="form-check-input" type="checkbox" id="showTotal" checked onchange="toggleTotal(this.checked)">
        </div>
      </div>
    </div>
    <div class="chart-tall"><canvas id="tsChart"></canvas></div>
  </div>
  <div class="bg-white rounded shadow-sm p-3">
    <h6 class="fw-bold mb-2">データ一覧 (最新30件)</h6>
    <div class="table-responsive" style="max-height:280px;overflow-y:auto;">
      <table class="table table-sm table-striped small" id="ts-table">
        <thead class="table-dark sticky-top"><tr id="ts-hdr"></tr></thead>
        <tbody id="ts-body"></tbody>
      </table>
    </div>
  </div>
</div>

<!-- ② サマリー -->
<div class="tab-pane fade" id="tab-sum">
  <div class="row g-3 mb-3" id="kpi-row"></div>
  <div class="row g-3">
    <div class="col-lg-5">
      <div class="bg-white rounded shadow-sm p-3">
        <div class="sec-header">資産クラス別 時価総額 (百万円)</div>
        <div class="chart-md"><canvas id="classChart"></canvas></div>
      </div>
    </div>
    <div class="col-lg-4">
      <div class="bg-white rounded shadow-sm p-3 mb-3">
        <div class="sec-header">USD / 円建て 比率</div>
        <div style="height:200px"><canvas id="currChart"></canvas></div>
      </div>
      <div class="bg-white rounded shadow-sm p-3">
        <div class="sec-header">人別 資産比率</div>
        <div style="height:200px"><canvas id="persChart"></canvas></div>
      </div>
    </div>
    <div class="col-lg-3">
      <div class="bg-white rounded shadow-sm p-3 h-100">
        <h6 class="fw-bold">資産クラス別一覧</h6>
        <table class="table table-sm small" id="cls-tbl">
          <thead class="table-dark"><tr><th>クラス</th><th class="text-end">百万円</th><th class="text-end">比率</th></tr></thead>
          <tbody id="cls-body"></tbody>
        </table>
        <h6 class="fw-bold mt-3">人別合計</h6>
        <table class="table table-sm small">
          <thead class="table-dark"><tr><th>名前</th><th class="text-end">百万円</th><th class="text-end">比率</th></tr></thead>
          <tbody id="per-body"></tbody>
        </table>
      </div>
    </div>
  </div>
</div>

<!-- ③ 資産クラス詳細 -->
<div class="tab-pane fade" id="tab-det">
  <ul class="nav nav-pills mb-3 flex-wrap small" id="secTab">{nav_html}</ul>
  <div class="tab-content bg-white rounded shadow-sm p-3" id="secTabContent">{content_html}</div>
</div>

</div><!-- tab-content -->
</div><!-- container -->

<div id="toast-box"></div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
<script>
// ── data ──────────────────────────────────────────────────
const ALL_DATES    = {dates_j};
const DS_MAIN      = {ds_main_j};
const DS_TOTAL     = {ds_tot_j};
const CLS_LABELS   = {cls_lbl_j};
const CLS_VALUES   = {cls_val_j};
const TOT_LABELS   = {tot_lbl_j};
const TOT_VALUES   = {tot_val_j};
const PERS_LABELS  = {pers_lbl_j};
const PERS_VALUES  = {pers_val_j};
const UDLABELS     = {usd_jpy_lbl};
const UDVALUES     = {usd_jpy_val};
const LIQUID_TOTAL = {liquid_total_j};  // 流動資産 = Total(不動産除く)
const TOTAL_INCL   = {total_incl_j};   // 総資産   = Total(不動産含む)

// ── 時系列チャート ─────────────────────────────────────
const allDS = [...DS_MAIN, ...DS_TOTAL.map(d=>Object.assign({{}},d,{{borderWidth:2,fill:false,borderDash:[5,3],pointRadius:0}}))];
const tsCtx = document.getElementById('tsChart');
const tsChart = new Chart(tsCtx, {{
  type:'line',
  data:{{labels:ALL_DATES, datasets:allDS}},
  options:{{
    responsive:true, maintainAspectRatio:false,
    interaction:{{mode:'index',intersect:false}},
    plugins:{{
      legend:{{position:'top',labels:{{boxWidth:10,font:{{size:10}}}}}},
      tooltip:{{callbacks:{{label:c=>c.dataset.label+': '+(c.raw!==null?c.raw.toFixed(1)+'M':'—')}}}}
    }},
    scales:{{
      x:{{type:'time',time:{{unit:'year'}},ticks:{{font:{{size:10}}}}}},
      y:{{ticks:{{callback:v=>v+'M',font:{{size:10}}}}}}
    }}
  }}
}});

function setRange(r){{
  const last = new Date(ALL_DATES[ALL_DATES.length-1]);
  let from = null;
  if(r==='1m'){{from=new Date(last);from.setMonth(from.getMonth()-1);}}
  else if(r==='6m'){{from=new Date(last);from.setMonth(from.getMonth()-6);}}
  else if(r==='1y'){{from=new Date(last);from.setFullYear(from.getFullYear()-1);}}
  else if(r==='3y'){{from=new Date(last);from.setFullYear(from.getFullYear()-3);}}
  else if(r==='5y'){{from=new Date(last);from.setFullYear(from.getFullYear()-5);}}
  tsChart.options.scales.x.min = from?from.toISOString().slice(0,10):undefined;
  tsChart.options.scales.x.max = last.toISOString().slice(0,10);
  tsChart.update();
}}

function toggleTotal(show){{
  const n = DS_MAIN.length;
  tsChart.data.datasets.forEach((d,i)=>{{ if(i>=n) d.hidden=!show; }});
  tsChart.update();
}}

// テーブル
(function(){{
  const hdr=document.getElementById('ts-hdr');
  const bod=document.getElementById('ts-body');
  const allLbls=['日付',...DS_MAIN.map(d=>d.label),...DS_TOTAL.map(d=>d.label)];
  hdr.innerHTML=allLbls.map(l=>`<th>${{l}}</th>`).join('');
  const n=ALL_DATES.length;
  for(let i=n-1;i>=Math.max(0,n-30);i--){{
    const cells=[ALL_DATES[i],...DS_MAIN.map(d=>d.data[i]!==null?d.data[i].toFixed(1):'—'),...DS_TOTAL.map(d=>d.data[i]!==null?d.data[i].toFixed(1):'—')];
    const tr=document.createElement('tr');
    tr.innerHTML=cells.map(c=>`<td>${{c}}</td>`).join('');
    bod.appendChild(tr);
  }}
}})();

// ── サマリー KPI ──────────────────────────────────────
(function(){{
  const row=document.getElementById('kpi-row');
  [
    ['流動資産 (不動産除く)', LIQUID_TOTAL.toFixed(1)+'M円'],
    ['USD建て資産', UDVALUES[0].toFixed(1)+'M円'],
    ['円建て資産',  UDVALUES[1].toFixed(1)+'M円'],
    ['USDJPY', '{usdjpy:.2f}'],
  ].forEach(([lbl,val])=>{{
    row.innerHTML+=`<div class="col-6 col-md-3"><div class="kpi-card"><div class="kpi-val">${{val}}</div><div class="kpi-lbl">${{lbl}}</div></div></div>`;
  }});
}})();

// 資産クラス棒グラフ
new Chart(document.getElementById('classChart'),{{
  type:'bar',
  data:{{labels:CLS_LABELS,datasets:[{{label:'百万円',data:CLS_VALUES,
    backgroundColor:['#4e79a7','#f28e2b','#e15759','#76b7b2','#59a14f','#edc948','#b07aa1','#ff9da7','#9c755f','#bab0ac']}}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}}}},
    scales:{{y:{{ticks:{{callback:v=>v+'M'}}}}}}
  }}
}});

// USD/円建て ドーナツ (流動資産ベース)
new Chart(document.getElementById('currChart'),{{
  type:'doughnut',
  data:{{labels:UDLABELS,datasets:[{{data:UDVALUES,backgroundColor:['#4e79a7','#e15759']}}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'bottom',labels:{{font:{{size:10}}}}}},
      tooltip:{{callbacks:{{label:c=>c.label+': '+c.raw.toFixed(1)+'M ('+(LIQUID_TOTAL>0?(c.raw/LIQUID_TOTAL*100).toFixed(1):'—')+'%)'}}}}
    }}
  }}
}});

// 人別ドーナツ (流動資産ベース)
new Chart(document.getElementById('persChart'),{{
  type:'doughnut',
  data:{{labels:PERS_LABELS,datasets:[{{data:PERS_VALUES,backgroundColor:['#4e79a7','#f28e2b','#e15759','#59a14f']}}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'bottom',labels:{{font:{{size:10}}}}}},
      tooltip:{{callbacks:{{label:c=>c.label+': '+c.raw.toFixed(1)+'M ('+(LIQUID_TOTAL>0?(c.raw/LIQUID_TOTAL*100).toFixed(1):'—')+'%)'}}}}
    }}
  }}
}});

// 資産クラステーブル (比率は流動資産ベース)
(function(){{
  const tb=document.getElementById('cls-body');
  CLS_LABELS.forEach((l,i)=>{{
    const v=CLS_VALUES[i]||0;
    tb.innerHTML+=`<tr><td>${{l}}</td><td class="text-end">${{v.toFixed(1)}}</td><td class="text-end">${{LIQUID_TOTAL>0?(v/LIQUID_TOTAL*100).toFixed(1):'—'}}%</td></tr>`;
  }});
  tb.innerHTML+=`<tr class="table-dark fw-bold"><td>流動資産合計</td><td class="text-end">${{LIQUID_TOTAL.toFixed(1)}}</td><td class="text-end">100%</td></tr>`;
  tb.innerHTML+=`<tr class="table-secondary fw-bold"><td>総資産（不動産含む）</td><td class="text-end">${{TOTAL_INCL.toFixed(1)}}</td><td class="text-end">—</td></tr>`;
}})();

// 人別テーブル (比率は流動資産(不動産除く)ベース)
(function(){{
  const tb=document.getElementById('per-body');
  let persSum=0;
  PERS_LABELS.forEach((l,i)=>{{
    const v=PERS_VALUES[i]||0;
    persSum+=v;
    tb.innerHTML+=`<tr><td>${{l}}</td><td class="text-end">${{v.toFixed(1)}}</td><td class="text-end fw-semibold">${{LIQUID_TOTAL>0?(v/LIQUID_TOTAL*100).toFixed(1):'—'}}%</td></tr>`;
  }});
  tb.innerHTML+=`<tr class="table-dark fw-bold"><td>合計</td><td class="text-end">${{persSum.toFixed(1)}}</td><td class="text-end">${{LIQUID_TOTAL>0?(persSum/LIQUID_TOTAL*100).toFixed(1):'—'}}%</td></tr>`;
}})();

// ── 編集機能 ──────────────────────────────────────────
function startEdit(btn){{
  const tr=btn.closest('tr');
  tr.querySelectorAll('td:not(:last-child)').forEach(td=>{{
    td.dataset.orig=td.textContent;
    const inp=document.createElement('input');
    inp.value=td.textContent.trim();
    inp.className='form-control form-control-sm p-0 px-1';
    td.innerHTML=''; td.appendChild(inp);
    td.classList.add('editing');
  }});
  btn.classList.add('d-none');
  tr.querySelector('.save-btn').classList.remove('d-none');
  tr.querySelector('.cancel-btn').classList.remove('d-none');
}}
function cancelEdit(btn){{
  const tr=btn.closest('tr');
  tr.querySelectorAll('td.editing').forEach(td=>{{
    td.textContent=td.dataset.orig||'';
    td.classList.remove('editing');
  }});
  btn.classList.add('d-none');
  tr.querySelector('.save-btn').classList.add('d-none');
  tr.querySelector('.edit-btn').classList.remove('d-none');
}}
function saveEdit(btn){{
  const tr=btn.closest('tr');
  tr.querySelectorAll('td.editing').forEach(td=>{{
    const inp=td.querySelector('input');
    if(inp){{td.textContent=inp.value;td.classList.remove('editing');}}
  }});
  btn.classList.add('d-none');
  tr.querySelector('.cancel-btn').classList.add('d-none');
  tr.querySelector('.edit-btn').classList.remove('d-none');
  toast('行 '+tr.dataset.row+' を更新しました (ローカル表示のみ)');
}}
function toast(msg){{
  const d=document.createElement('div');
  d.className='toast align-items-center text-bg-success border-0 show mb-2';
  d.innerHTML=`<div class="d-flex"><div class="toast-body small">${{msg}}</div><button type="button" class="btn-close btn-close-white me-2 m-auto" onclick="this.closest('.toast').remove()"></button></div>`;
  document.getElementById('toast-box').appendChild(d);
  setTimeout(()=>d.remove(),3000);
}}
function exportSection(k){{
  const t=document.getElementById('table-'+k);
  if(!t)return;
  const rows=t.querySelectorAll('tr');
  const csv=[...rows].map(r=>[...r.querySelectorAll('th,td')].slice(0,-1).map(c=>'"'+c.textContent.trim()+'"').join(',')).join('\\n');
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob(['\\uFEFF'+csv],{{type:'text/csv;charset=utf-8'}}));
  a.download='section_'+k+'.csv'; a.click();
}}
</script>
</body>
</html>'''

with open(OUT_HTML, 'w', encoding='utf-8') as f:
    f.write(html)
print(f'HTML保存: {OUT_HTML}')
print('完了!')
